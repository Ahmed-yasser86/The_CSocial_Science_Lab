"""FastAPI application for the SocialScienceResearch module.

Mounts the research API under the configured prefix (default
``/api/v1/social-science``). Endpoints map onto the service layer
(collection, sampling, analytics, recommendation network) and depend only on
the service interfaces.

The application is created via :func:`create_app`, which wires services to a
single persistence store so collection and analytics read the same data.

API hardening (B2)
------------------
* every endpoint declares a pydantic ``response_model`` (``api.schemas``);
* list endpoints use opaque cursor pagination (``services.pagination``);
* domain errors map to a single machine-readable error envelope;
* CORS origins, OpenAPI metadata and docs visibility come from settings.
"""

from __future__ import annotations

import asyncio
import json
from contextlib import asynccontextmanager
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Annotated, Any
from io import BytesIO

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from pydantic import BaseModel, Field

from SocialScienceResearch.config.settings import SocialScienceSettings
from SocialScienceResearch.domain.collection import CollectionSpec
from SocialScienceResearch.domain.enums import RunType
from SocialScienceResearch.domain.layer_models import LayerRun
from SocialScienceResearch.domain.query import (
    OPERATOR_DESCRIPTIONS,
    ResearchQueryRequest,
    SamplingSpec,
    AdvancedSamplingSpec,
    VideoFilter,
    evaluate_query,
    preview_query,
)
from SocialScienceResearch.persistence.factory import build_repositories
from SocialScienceResearch.services import (
    AnalyticsService,
    CoverageReport,
    JobManager,
    QualityService,
    QueryService,
    RecommendationGraphService,
    RecommendationService,
    SamplingService,
)
from SocialScienceResearch.services.layer_scrape_service import LayerScrapeService
from SocialScienceResearch.services.echo_chamber_service import EchoChamberService
from SocialScienceResearch.services.pagination import (
    CursorError,
    Paginated,
    page_sorted,
)
from SocialScienceResearch.services.sampling_service import SamplingError
from SocialScienceResearch.services.variable_registry import VariableRegistry

from .schemas import (
    ChannelCountPayload,
    ChannelOverviewPayload,
    ChannelPayload,
    CollectionErrorPayload,
    CollectionResultPayload,
    CollectionResultsPayload,
    CommentPayload,
    CommentStatsPayload,
    DatasetSummaryPayload,
    ErrorPayload,
    ExportRequest,
    ExportResponse,
    JobCancelPayload,
    JobFailurePayload,
    JobPayload,
    JobResultPayload,
    JobSubmitPayload,
    NetworkSummaryPayload,
    OperatorInfoPayload,
    PercentilesPayload,
    QueryPreviewResponse,
    QueryResolveResponse,
    RawVideoPayload,
    RecommendationPayload,
    RunPayload,
    RunVideosPayload,
    SamplingResultPayload,
    SystemFoldersPayload,
    ThreadPayload,
    TopVideosPayload,
    TopVideoRow,
    UpdateRunRequest,
    JobTagsRequest,
    VariableMetaPayload,
    VelocityPoint,
    VideoEngagementPayload,
    VideoNetworkContextPayload,
    VideoObservationPayload,
    VideoPayload,
)

#: Default page size for cursor-paginated list endpoints.
DEFAULT_PAGE_SIZE = 50


class CollectRequest(BaseModel):
    url: str


class ScrapeRecommendationsRequest(BaseModel):
    video_url: str


class NetworkScrapeVideoRequest(BaseModel):
    """Click-to-scrape one video node from the network tab."""

    video_id: str
    trigger_run_id: str | None = None
    tags: list[str] = Field(default_factory=list)


class NetworkScrapeRunRequest(BaseModel):
    """Bulk re-scrape of every source video observed in one run."""

    run_id: str
    dedupe: bool = True
    tags: list[str] = Field(default_factory=list)


class NetworkScrapeChannelRequest(BaseModel):
    """Bulk scrape of every known video belonging to a channel."""

    channel_id: str
    trigger_run_id: str | None = None
    dedupe: bool = True
    tags: list[str] = Field(default_factory=list)


def build_services(
    settings: SocialScienceSettings,
    *,
    provider=None,
    repository=None,
    jobs_manager: "JobManager | None" = None,
) -> dict[str, Any]:
    """Build the full service container for one persistence binding.

    ``repository`` overrides ``settings.repository`` so the SAME factory
    provisions services against a workspace's database + data dir (plan 2.3).
    ``jobs_manager`` reuses an existing JobManager across workspace switches:
    the manager owns the live job registry and its executor queue, so every
    rebuild MUST share one instance - otherwise services like the echo
    detector submit to a manager that no longer serves reads, and the merged
    view reconciles genuinely-live jobs to 'interrupted'.
    """
    repo_settings = repository or settings.repository
    repos = build_repositories(repo_settings)
    if provider is None:
        from SocialScienceResearch.acquisition import YtDlpAcquisitionProvider

        provider = YtDlpAcquisitionProvider(
            settings=settings.scraper, collection=settings.collection
        )
    # Crash honesty (plan J1): a job that was pending/running when the
    # previous process died can never finish; mark those rows interrupted.
    try:
        repos.jobs.reconcile_stale_running(
            "interrupted: the server restarted while this job was active"
        )
    except Exception:  # noqa: BLE001 - persistence is best-effort at boot
        pass
    if jobs_manager is None:
        jobs_manager = JobManager(
            max_workers=settings.jobs.max_workers,
            max_run_seconds=settings.jobs.max_run_seconds,
            max_stall_seconds=settings.jobs.max_stall_seconds,
            store=repos.jobs,
        )
    else:
        # Same live registry, new workspace's write-through store.
        jobs_manager.set_store(repos.jobs)
    return {
        "repos": repos,
        # ``RecommendationService`` extends ``CollectionService``; using it for
        # the ``"collection"`` key wires the spec-driven recommendation target
        # (previously raised NotImplementedError from the base class).
        "collection": RecommendationService(provider, repos, settings=settings),
        "recommendations": RecommendationService(provider, repos, settings=settings),
        "analytics": AnalyticsService(repos),
        "query": QueryService(repos, settings),
        "sampling": SamplingService(repos, settings.sampling.default_seed),
        "network": RecommendationGraphService(repos),
        "quality": QualityService(repos),
        "jobs": jobs_manager,
        "layer_scrape": LayerScrapeService(provider, repos, settings=settings),
        "echo": EchoChamberService(
            provider, repos, settings=settings, jobs=jobs_manager
        ),
    }


#: Keys owned by :func:`build_services`; anything else found in the live
#: container was lazily cached there by a router (``common.get_service``) and
#: must be dropped on a workspace switch because it closes over stale repos.
_CORE_SERVICE_KEYS = frozenset({
    "repos",
    "collection",
    "recommendations",
    "analytics",
    "query",
    "sampling",
    "network",
    "quality",
    "jobs",
    "layer_scrape",
    "echo",
})


class _ActiveRepos:
    """Attribute-forwarding view of the CURRENT workspace's repositories.

    ``create_app`` exposes ``repos = _ActiveRepos(services)`` to its direct
    routes, so every ``repos.videos.list_videos(...)`` call site resolves
    against whichever repository container the active workspace routing last
    installed - no per-route churn, no stale pool access after a switch.
    """

    __slots__ = ("_services",)

    def __init__(self, services: dict[str, Any]) -> None:
        self._services = services

    def __getattr__(self, name: str) -> Any:
        return getattr(self._services["repos"], name)


class WorkspaceRuntime:
    """Connection routing: binds request handling to the ACTIVE workspace.

    The active pointer lives in ``<root data_dir>/workspaces/active.json``
    (server-side authoritative state). :meth:`sync` compares it with the
    currently-bound workspace and, on a switch, rebuilds every persistence-
    bound service in place (the SAME dict object is mutated so both the
    closures in :func:`create_app` and the routers reading
    ``request.app.state.services`` observe the new binding), clears the
    class-level graph/overlap caches and disposes the previous pool.
    """

    def __init__(self, settings: SocialScienceSettings, *, provider=None) -> None:
        self.settings = settings
        self.provider = provider
        from SocialScienceResearch.services.workspace_service import WorkspaceService

        self.workspaces = WorkspaceService(settings)
        # First run registers the existing default DB/data dir as the
        # renamable Legacy workspace; subsequent runs are no-ops.
        self.workspaces.bootstrap()
        self.services: dict[str, Any] = {}
        # Signature of the persistence binding currently loaded in
        # ``self.services``; starts as the default (Legacy) configuration.
        self.active_workspace_id: str | None = None
        self._bound: tuple[str, str] = (
            settings.repository.database_url,
            str(Path(settings.repository.data_dir)),
        )
        self._runtime_config: Any = None

    def attach_runtime_config(self, runtime_config: Any) -> None:
        """Remember the mutable scraper config for post-switch rewiring."""
        self._runtime_config = runtime_config

    def sync(self, app: FastAPI) -> None:
        """Ensure ``self.services`` is bound to the active workspace."""
        try:
            workspace_id = self.workspaces.active_workspace_id()
            if workspace_id is not None:
                self.workspaces.get(workspace_id)  # dangling-pointer check
            else:
                workspace_id = None
        except KeyError:
            # Registry lost the pointed-at workspace: deactivate instead of
            # serving requests against an unknown database.
            self.workspaces.deactivate()
            workspace_id = None

        if workspace_id == self.active_workspace_id and self._bound is not None:
            return

        workspace = (
            self.workspaces.get(workspace_id) if workspace_id is not None else None
        )
        repository = (
            self.workspaces.repository_settings(workspace)
            if workspace is not None
            else None
        )
        if workspace is not None:
            signature: tuple[str, str] | None = (
                workspace.database_url,
                str(Path(workspace.data_dir)),
            )
        else:
            signature = None
        if (
            signature is not None
            and signature == self._bound
            and self.services
        ):
            # Pointer already matches the loaded binding (e.g. Legacy).
            self.active_workspace_id = workspace_id
            return

        old_repos = self.services.get("repos")
        # The JobManager survives activation by design (its queue must be
        # empty - activation is guarded while jobs are pending/running). Pass
        # the SURVIVING manager into the rebuild so every freshly built
        # service (echo detector, layer crawl, ...) shares the ONE live job
        # registry; its store is rebound to the new workspace inside
        # build_services. Building a second manager here would make submits
        # land in a registry that read paths never see.
        fresh = build_services(
            self.settings,
            provider=self.provider,
            repository=repository,
            jobs_manager=self.services.get("jobs"),
        )
        for key in list(self.services):
            if key not in _CORE_SERVICE_KEYS:
                del self.services[key]
        self.services.clear()
        self.services.update(fresh)
        if self._runtime_config is not None:
            self.services["layer_scrape"].set_runtime_config(self._runtime_config)
            self.services["recommendations"].set_runtime_config(self._runtime_config)
        # Class-level caches have no workspace dimension: clear them on EVERY
        # switch (pitfalls R1/A1 defense-in-depth; instance TTL caches die
        # with the discarded service objects).
        RecommendationGraphService.clear_graph_cache()
        from SocialScienceResearch.services.commenter_overlap_service import (
            CommenterOverlapService,
        )

        CommenterOverlapService.clear_overlap_cache()
        if old_repos is not None and old_repos is not self.services["repos"]:
            old_repos.store.close()
        self.active_workspace_id = workspace_id
        self._bound = signature

    def active_jobs(self) -> list[Any]:
        """Jobs still pending/running per the ONE authoritative merge rule.

        A job is ACTIVE iff the live manager says pending/running; otherwise
        its persisted row governs. Persisted pending/running rows without a
        live owner were written by a previous process lifetime and are lazily
        finalised as ``interrupted`` here, so a restart can never leave a
        phantom blocking workspace activation forever.
        """
        manager = self.services.get("jobs")
        if manager is None:
            return []
        active = []
        for job in manager.list():
            status = getattr(job.status, "value", str(job.status))
            if status in ("pending", "running"):
                active.append(job)
        repos = self.services.get("repos")
        if repos is not None:
            for status_filter in _STALE_JOB_STATUSES:
                try:
                    rows = repos.jobs.list_jobs(status=status_filter)
                except Exception:  # noqa: BLE001 - healing is best-effort
                    continue
                for row in rows:
                    _reconcile_stale_job_row(repos, manager, row)
        return active


# Backwards-compatible alias (module was referenced as ``_services`` before the
# workspace-routing refactor; nothing outside this module should use it).
_services = build_services


def _run_key(run) -> tuple[str, ...]:
    return (run.started_at.isoformat(), run.run_id)


def _job_key(job) -> tuple[str, ...]:
    created = getattr(job, "created_at", None)
    created_key = created.isoformat() if isinstance(created, datetime) else str(created)
    return (created_key, getattr(job, "job_id", ""))


def _job_payload(job, runs: list[Any] | None = None) -> dict[str, Any]:
    """Best-effort JSON-safe payload for a single job.

    A single malformed job must never take down the whole list, so any
    serialization failure falls back to a minimal but valid payload. ``runs``
    carries the nested run summaries (plan J1 children provenance).
    """
    run_summaries = [_run_summary(r) for r in (runs or [])]
    try:
        payload = JobPayload.model_validate(job.to_dict()).model_dump(mode="json")
        payload["runs"] = run_summaries
        return payload
    except Exception:
        created = getattr(job, "created_at", None)
        created_iso = created.isoformat() if isinstance(created, datetime) else None
        return {
            "job_id": getattr(job, "job_id", "unknown"),
            "kind": getattr(job, "kind", ""),
            "status": str(getattr(job, "status", "")),
            "created_at": created_iso or datetime.now(timezone.utc).isoformat(),
            "cancel_requested": bool(getattr(job, "cancel_requested", False)),
            "runs": run_summaries,
        }


def _run_summary(run) -> dict[str, Any]:
    """Compact nested run row for a job's children table (additive field)."""
    return {
        "run_id": run.run_id,
        "run_type": run.run_type.value if hasattr(run.run_type, "value") else str(run.run_type),
        "target_url": run.target_url,
        "target_video_id": run.target_video_id,
        "parent_run_id": run.parent_run_id,
        "layer_index": run.layer_index,
        "status": (
            run.status.value if hasattr(run.status, "value") else str(run.status)
        ),
        "started_at": run.started_at.isoformat() if run.started_at else None,
        "finished_at": run.finished_at.isoformat() if run.finished_at else None,
        "entities_discovered": run.entities_discovered,
        "entities_succeeded": run.entities_succeeded,
        "entities_failed": run.entities_failed,
        "comments_collected": run.comments_collected,
        "name": run.name,
    }


def _runs_for_job(repos, job_id: str) -> list[Any]:
    """All runs stamped with this job id (nested children, plan J1)."""
    return [run for run in repos.runs.list_runs() if run.job_id == job_id]


#: Statuses a persisted row can carry that only make sense while a worker
#: thread of SOME process owns the job. The authoritative activity rule is:
#: a job is ACTIVE iff the live manager says pending/running; otherwise its
#: persisted row governs. A row left pending/running without a live owner was
#: written by a previous process lifetime (crash/restart lost the race with
#: boot-time ``reconcile_stale_running``) and is lazily finalised here.
_STALE_JOB_STATUSES = ("pending", "running")


def _reconcile_stale_job_row(repos, manager, persisted) -> bool:
    """Finalise an orphaned pending/running persisted row; True if it did.

    Marks the row ``interrupted`` so it never resurfaces as a phantom
    cancellable/running job in the merged view (and never blocks a workspace
    switch). Rows owned by a live worker are left untouched.
    """
    status = str(persisted.status or "")
    if status not in _STALE_JOB_STATUSES or manager.get(persisted.job_id) is not None:
        return False
    try:
        persisted.status = "interrupted"
        persisted.error = (
            "interrupted: no worker owns this job "
            "(written by a previous server process)"
        )
        persisted.message = "interrupted"
        persisted.finished_at = datetime.now(timezone.utc)
        persisted.updated_at = persisted.finished_at
        repos.jobs.save_job(persisted)
    except Exception:  # noqa: BLE001 - reconciliation is best-effort
        return False
    return True


def _persisted_job_row(repos, manager, job_id: str):
    """Fetch one persisted job row, lazily reconciling an orphaned row.

    Returns ``(row, reconciled)``; ``row`` is ``None`` for unknown ids.
    """
    try:
        persisted = repos.jobs.get_job(job_id)
    except Exception:  # noqa: BLE001 - persistence hiccups degrade to None
        return None, False
    if persisted is None:
        return None, False
    return persisted, _reconcile_stale_job_row(repos, manager, persisted)


def _job_from_persisted(persisted):
    """Adapt a persisted :class:`CollectionJob` row to the live Job shape."""

    class _PersistedJobView:
        def __init__(self, row):
            self.job_id = row.job_id
            self.kind = row.kind
            self.status = row.status
            self.created_at = row.created_at
            self.started_at = row.started_at
            self.finished_at = row.finished_at
            self.progress = {}
            self.message = row.message or row.error
            self.cancel_requested = False

        def to_dict(self) -> dict[str, Any]:
            """JSON-safe snapshot mirroring the live ``Job.to_dict`` shape."""
            return {
                "job_id": self.job_id,
                "kind": self.kind,
                "status": str(self.status or ""),
                "created_at": (
                    self.created_at.isoformat()
                    if isinstance(self.created_at, datetime)
                    else None
                ),
                "started_at": (
                    self.started_at.isoformat()
                    if isinstance(self.started_at, datetime)
                    else None
                ),
                "finished_at": (
                    self.finished_at.isoformat()
                    if isinstance(self.finished_at, datetime)
                    else None
                ),
                "progress": {},
                "message": self.message,
                "cancel_requested": False,
            }

    return _PersistedJobView(persisted)


def _job_dict_key(job_dict: dict[str, Any]) -> tuple[str, ...]:
    """Sort key for merged job payload dicts (created_at ISO then id)."""
    return ((job_dict.get("created_at") or ""), job_dict.get("job_id", ""))


def _video_key(video) -> tuple[str, ...]:
    return (video.video_id,)


def _channel_key(channel) -> tuple[str, ...]:
    return (channel.channel_id,)


def _obs_key(obs) -> tuple[str, ...]:
    return (obs.observed_at.isoformat(), obs.observation_id)


def _comment_key(comment) -> tuple[str, ...]:
    return (comment.comment_id,)


def _recommendation_edge_key(edge) -> tuple[str, ...]:
    """Sort key for recommendation edges: feed rank (position) then identity.

    Positions are zero-padded so string comparison mirrors numeric ordering,
    and ``None`` positions sort last (unknown rank). All keys are strings so
    cursor tokens remain comparable inside ``page_sorted``.
    """
    if isinstance(edge, dict):
        position = edge.get("position")
        position_key = f"{position:08d}" if position is not None else "~"
        return (
            position_key,
            edge.get("collection_run_id") or "",
            edge.get("observation_id", ""),
        )
    position = edge.position
    position_key = f"{position:08d}" if position is not None else "~"
    return (
        position_key,
        edge.collection_run_id or "",
        edge.observation_id,
    )


def _paginate(
    entities: list, *, cursor: str | None, page_size: int, key, reverse: bool = False
) -> Paginated[Any]:
    """Slice a materialized entity list into a ``Paginated`` envelope.

    ``total`` is always populated because the repositories return in-memory
    lists (research scale), making the count free. ``reverse`` yields pages
    newest-first (see ``page_sorted``).
    """
    full = sorted(entities, key=key)
    page = page_sorted(
        full,
        cursor=cursor,
        page_size=page_size,
        key_func=key,
        total=len(full),
        reverse=reverse,
    )
    items = []
    for e in page.items:
        if isinstance(e, dict):
            items.append(e)
        elif hasattr(e, 'model_dump'):
            items.append(e.model_dump())
        else:
            items.append(e.__dict__)
    return Paginated(
        items=items,
        next_cursor=page.next_cursor,
        has_more=page.has_more,
        total=page.total,
    )


def create_app(
    settings: SocialScienceSettings | None = None, *, provider=None
) -> FastAPI:
    settings = settings or SocialScienceSettings()
    runtime = WorkspaceRuntime(settings, provider=provider)
    services = runtime.services
    services.update(build_services(settings, provider=provider))
    runtime.active_workspace_id = None  # synced to the persisted pointer below

    @asynccontextmanager
    async def lifespan(app: FastAPI):
        # Pre-warm the graph cache in the background so the first user request
        # doesn't block on a cold ~20s rebuild (which the proxy times out on).
        import asyncio as _asyncio
        def _warm() -> None:
            try:
                from SocialScienceResearch.services.recommendation_graph_service import (
                    RecommendationGraphService,
                )
                RecommendationGraphService(services["repos"]).build_graph(run_id=None)
            except Exception:  # noqa: BLE001
                pass
        _asyncio.get_event_loop().run_in_executor(None, _warm)
        yield
        services["repos"].store.close()
        services["jobs"].shutdown()

    app = FastAPI(
        title=settings.api.title,
        version=settings.api.version,
        description=settings.api.description,
        docs_url="/docs" if settings.api.docs_enabled else None,
        redoc_url="/redoc" if settings.api.docs_enabled else None,
        lifespan=lifespan,
    )
    app.state.services = services
    app.state.settings = settings
    app.state.workspace_runtime = runtime

    # Request-scoped workspace routing: every request is served by the
    # services bound to the ACTIVE workspace's database + data dir. Installed
    # as HTTP middleware so both these direct routes (whose closures share the
    # mutated ``services`` dict) and every router reading
    # ``request.app.state.services`` are covered without per-route churn.
    @app.middleware("http")
    async def route_to_active_workspace(request: Request, call_next):
        runtime.sync(request.app)
        return await call_next(request)

    # Mutable runtime scraper config (UI can update without restart).
    from SocialScienceResearch.config.runtime_config import RuntimeScraperConfig
    app.state.runtime_scraper_config = RuntimeScraperConfig(
        request_delay_seconds=settings.scraper.request_delay_seconds,
        enrichment_concurrency=settings.scraper.enrichment_concurrency,
        socket_timeout=settings.scraper.socket_timeout,
        retries=settings.scraper.retries,
        retry_backoff=settings.scraper.retry_backoff,
    )
    # Wire runtime config into services so they read mutable settings
    # instead of frozen ones.
    services["layer_scrape"].set_runtime_config(app.state.runtime_scraper_config)
    services["recommendations"].set_runtime_config(app.state.runtime_scraper_config)

    app.add_middleware(
        CORSMiddleware,
        allow_origins=settings.api.cors_origins,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    # ------------------------------------------------------------------
    # Error envelope: domain errors map to machine-readable codes.
    # ------------------------------------------------------------------
    @app.exception_handler(HTTPException)
    def _http_exception_handler(request: Request, exc: HTTPException) -> JSONResponse:
        return JSONResponse(
            status_code=exc.status_code,
            content=ErrorPayload(
                code=f"http_{exc.status_code}", message=str(exc.detail)
            ).model_dump(),
        )

    @app.exception_handler(SamplingError)
    def _sampling_exception_handler(request: Request, exc: SamplingError) -> JSONResponse:
        return JSONResponse(
            status_code=400,
            content=ErrorPayload(code="sampling_error", message=str(exc)).model_dump(),
        )

    @app.exception_handler(CursorError)
    def _cursor_exception_handler(request: Request, exc: CursorError) -> JSONResponse:
        return JSONResponse(
            status_code=400,
            content=ErrorPayload(code="invalid_cursor", message=str(exc)).model_dump(),
        )

    @app.exception_handler(ValueError)
    def _value_error_handler(request: Request, exc: ValueError) -> JSONResponse:
        return JSONResponse(
            status_code=400,
            content=ErrorPayload(code="invalid_argument", message=str(exc)).model_dump(),
        )

    @app.exception_handler(KeyError)
    def _key_error_handler(request: Request, exc: KeyError) -> JSONResponse:
        # Services raise KeyError for an unknown entity id (same convention as
        # the commenters profile endpoint); bad *input* stays a ValueError.
        message = str(exc.args[0]) if exc.args else str(exc)
        return JSONResponse(
            status_code=404,
            content=ErrorPayload(code="not_found", message=message).model_dump(),
        )

    prefix = settings.api.prefix
    repos = _ActiveRepos(services)

    # ------------------------------------------------------------------
    # Phase B-D routers (split modules so they build in parallel).
    # Included BEFORE the direct routes below so literal paths such as
    # ``/runs/delta`` are never shadowed by ``/runs/{run_id}``.
    # ------------------------------------------------------------------
    from .routers import (
        channels,
        commenters,
        comments,
        comparison,
        content_homophily,
        datasets,
        echo_chamber,
        explorer,
        expansion,
        layer_network,
        network_ext,
        project_items,
        samples,
        scraper_config,
        search,
        session,
        workspaces,
    )

    app.include_router(channels.router, prefix=prefix)
    app.include_router(commenters.router, prefix=prefix)
    app.include_router(comments.router, prefix=prefix)
    app.include_router(comparison.router, prefix=prefix)
    app.include_router(content_homophily.router, prefix=prefix)
    app.include_router(datasets.router, prefix=prefix)
    app.include_router(echo_chamber.router, prefix=prefix)
    app.include_router(explorer.router, prefix=prefix)
    app.include_router(expansion.router, prefix=prefix)
    app.include_router(layer_network.router, prefix=prefix)
    app.include_router(network_ext.router, prefix=prefix)
    app.include_router(project_items.router, prefix=prefix)
    app.include_router(samples.router, prefix=prefix)
    app.include_router(scraper_config.router, prefix=prefix)
    app.include_router(search.router, prefix=prefix)
    app.include_router(session.router, prefix=prefix)
    app.include_router(workspaces.router, prefix=prefix)

    # Align the initial service binding with the persisted active-workspace
    # pointer (no-op when it still points at the default/Legacy configuration).
    runtime.sync(app)


    # ------------------------------------------------------------------
    # Collection
    # ------------------------------------------------------------------
    @app.post(
        f"{prefix}/collect/channel",
        tags=["collection"],
        response_model=CollectionResultPayload,
    )
    def collect_channel(body: CollectRequest):
        return _collection_payload(services["collection"].collect_channel(body.url))

    @app.post(
        f"{prefix}/collect/video",
        tags=["collection"],
        response_model=CollectionResultPayload,
    )
    def collect_video(body: CollectRequest):
        return _collection_payload(services["collection"].collect_video(body.url))

    @app.post(
        f"{prefix}/collect/recommendations",
        tags=["collection"],
        response_model=CollectionResultPayload,
    )
    def collect_recommendations(body: CollectRequest):
        return _collection_payload(
            services["recommendations"].collect_recommendations(body.url)
        )

    # ------------------------------------------------------------------
    # Spec-driven collection (async jobs with progress + cancellation)
    # ------------------------------------------------------------------
    @app.post(f"{prefix}/collect", tags=["collection"], response_model=JobSubmitPayload)
    def collect_spec(spec: CollectionSpec):
        """Submit a spec-driven collection experiment; returns a job id.

        Runs in the background (worker thread) so the client can poll progress
        via ``GET /jobs/{job_id}`` and cancel via ``POST /jobs/{job_id}/cancel``.
        """

        def _worker(reporter):
            return services["collection"].collect(spec, reporter=reporter)

        job = services["jobs"].submit(_worker, kind="collect")
        return {"job_id": job.job_id}

    @app.post(f"{prefix}/scrape/recommendations", tags=["collection"], response_model=JobSubmitPayload)
    def scrape_recommendations(body: ScrapeRecommendationsRequest):
        """Submit a recommendation scrape job for a single video.
        
        Runs in the background (worker thread) so the client can poll progress
        via ``GET /jobs/{job_id}`` and cancel via ``POST /jobs/{job_id}/cancel``.
        """
        def _worker(reporter):
            return services["recommendations"].collect_recommendations(body.video_url, reporter=reporter)

        job = services["jobs"].submit(_worker, kind="recommendation", tags=body.tags)
        return {"job_id": job.job_id}

    @app.post(
        f"{prefix}/network/scrape/video",
        tags=["network"],
        response_model=JobSubmitPayload,
    )
    def network_scrape_video(body: NetworkScrapeVideoRequest):
        """Queue a recommendation scrape for a single video (graph node click)."""
        def _worker(reporter):
            return services["recommendations"].collect_recommendations(
                video_url=None,
                video_id=body.video_id,
                parent_run_id=body.trigger_run_id,
                reporter=reporter,
            )

        job = services["jobs"].submit(_worker, kind="recommendation", tags=body.tags)
        return {"job_id": job.job_id}

    @app.post(
        f"{prefix}/network/scrape/run",
        tags=["network"],
        response_model=JobSubmitPayload,
    )
    def network_scrape_run(body: NetworkScrapeRunRequest):
        """Queue a bulk re-scrape of every source video observed in one run."""
        run = repos.runs.get_run(body.run_id)
        if run is None:
            raise HTTPException(status_code=404, detail=f"Run {body.run_id} not found")

        def _resolve_sources() -> list[str]:
            if run.run_type == RunType.CHANNEL:
                return [v.video_id for v in repos.videos.list_videos_by_run(run.run_id)]
            sources = sorted(
                {
                    e.source_video_id
                    for e in repos.recommendations.list_recommendation_edges(
                        run_id=run.run_id
                    )
                }
            )
            return sources

        def _worker(reporter):
            sources = _resolve_sources()
            if not sources:
                raise HTTPException(
                    status_code=400,
                    detail=f"Run {run.run_id} has no source videos to re-scrape",
                )
            return services["recommendations"].collect_recommendations_for_videos(
                sources,
                parent_run_id=run.run_id,
                dedupe_run_ids=[run.run_id] if body.dedupe else None,
                dedupe_all_history=body.dedupe,
                reporter=reporter,
            )

        job = services["jobs"].submit(_worker, kind="recommendation", tags=body.tags)
        return {"job_id": job.job_id}

    @app.post(
        f"{prefix}/network/scrape/channel",
        tags=["network"],
        response_model=JobSubmitPayload,
    )
    def network_scrape_channel(body: NetworkScrapeChannelRequest):
        """Queue a bulk scrape of every known video belonging to a channel."""
        def _worker(reporter):
            videos = repos.videos.list_videos(channel_id=body.channel_id)
            if not videos:
                raise HTTPException(
                    status_code=400,
                    detail=f"Channel {body.channel_id} has no persisted videos",
                )
            return services["recommendations"].collect_recommendations_for_videos(
                [v.video_id for v in videos],
                parent_run_id=body.trigger_run_id,
                channel_id=body.channel_id,
                dedupe_run_ids=[body.trigger_run_id] if (body.dedupe and body.trigger_run_id) else None,
                dedupe_all_history=body.dedupe,
                reporter=reporter,
            )

        job = services["jobs"].submit(_worker, kind="recommendation", tags=body.tags)
        return {"job_id": job.job_id}

    @app.get(f"{prefix}/jobs", tags=["jobs"], response_model=Paginated[JobPayload])
    def list_jobs(
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
        kind: str | None = Query(None, description="Filter by job kind"),
        status: str | None = Query(None, description="Filter by job status"),
        created_after: datetime | None = Query(None, description="Only jobs created after this timestamp"),
    ):
        # Live (in-memory) jobs take precedence; persisted rows not in memory
        # (older than the current process) are merged so completed jobs remain
        # listed across restarts (plan J1 US-J1).
        runs_by_job: dict[str, list[Any]] = {}
        for run in repos.runs.list_runs():
            if run.job_id:
                runs_by_job.setdefault(run.job_id, []).append(run)
        merged: dict[str, dict[str, Any]] = {}
        for persisted in services["jobs"].persisted_jobs(kind=kind, status=status):
            # Lazy reconciliation: a pending/running row with no live owner
            # is finalised before it can surface as a phantom active job.
            _reconcile_stale_job_row(services["repos"], services["jobs"], persisted)
            payload = _job_payload(
                _job_from_persisted(persisted), runs=runs_by_job.get(persisted.job_id, [])
            )
            payload["result_json"] = persisted.result_json or {}
            merged[persisted.job_id] = payload
        for live in services["jobs"].list():
            if kind and live.kind != kind:
                continue
            live_status = getattr(live.status, "value", str(live.status))
            if status and live_status != status:
                continue
            merged[live.job_id] = _job_payload(live, runs=runs_by_job.get(live.job_id, []))
        jobs_list = list(merged.values())
        if created_after is not None:
            threshold = created_after.isoformat()
            jobs_list = [
                j for j in jobs_list if (j.get("created_at") or "") >= threshold
            ]
        jobs_sorted = sorted(jobs_list, key=_job_dict_key)
        page = page_sorted(
            jobs_sorted,
            cursor=cursor,
            page_size=page_size,
            key_func=_job_dict_key,
            total=len(jobs_sorted),
        )
        return Paginated(
            items=page.items,
            next_cursor=page.next_cursor,
            has_more=page.has_more,
            total=page.total,
        )

    @app.post(
        f"{prefix}/jobs/kill-stuck",
        tags=["jobs"],
    )
    def kill_stuck_jobs():
        """Force-terminate every pending/running job and recycle the worker pool.

        Jobs blocked on a stalled yt-dlp/network call cannot be cancelled
        cooperatively (cancellation is only honoured at unit boundaries), so
        this marks them terminal and abandons their worker threads, unblocking
        the queue. Intended as an explicit operator escape hatch.

        Declared before ``/jobs/{job_id}`` so the static path wins over the
        ``{job_id}`` path parameter for POST requests.
        """
        return services["jobs"].kill_stuck()

    @app.get(f"{prefix}/jobs/{{job_id}}", tags=["jobs"], response_model=JobPayload)
    def get_job(job_id: str):
        job = services["jobs"].get(job_id)
        if job is not None:
            return _job_payload(job, runs=_runs_for_job(repos, job_id))
        # Not live anymore (process restart): serve the persisted row. A row
        # still claiming pending/running without a live owner is reconciled
        # first so the merged view never shows a phantom running job.
        persisted, _reconciled = _persisted_job_row(
            services["repos"], services["jobs"], job_id
        )
        if persisted is None:
            raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
        payload = _job_payload(
            _job_from_persisted(persisted), runs=_runs_for_job(repos, job_id)
        )
        payload["result_json"] = persisted.result_json or {}
        return payload

    @app.post(
        f"{prefix}/jobs/{{job_id}}/cancel",
        tags=["jobs"],
        response_model=JobCancelPayload,
    )
    def cancel_job(job_id: str):
        if services["jobs"].cancel(job_id):
            return {"job_id": job_id, "cancelled": True}
        # Not live: the persisted row governs. A pending/running row with no
        # live owner (previous process lifetime) is finalised as interrupted
        # by the lookup below, so honouring the cancellation succeeds instead
        # of 409-ing against a phantom the UI rightly showed as active.
        persisted, reconciled = _persisted_job_row(
            repos, services["jobs"], job_id
        )
        if reconciled:
            return {"job_id": job_id, "cancelled": True}
        detail = (
            f"Job {job_id} cannot be cancelled "
            f"(status: {persisted.status})"
            if persisted is not None
            else f"Job {job_id} cannot be cancelled (finished or missing)"
        )
        raise HTTPException(status_code=409, detail=detail)

    @app.get(
        f"{prefix}/jobs/{{job_id}}/result",
        tags=["jobs"],
        response_model=JobResultPayload,
    )
    def job_result(job_id: str):
        job = services["jobs"].get(job_id)
        if job is None:
            raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
        if job.status.value in ("pending", "running"):
            raise HTTPException(status_code=409, detail="Job is still running")
        if job.error:
            return {"error": job.error}
        result = job.result
        # Layer-crawl and network-expansion jobs store a ``LayerRun`` anchor as
        # their result (not a ``CollectionResult``), so the generic
        # collection-shape serializer below would raise ``AttributeError`` and
        # 500. Serialize those results explicitly and surface them as extra
        # fields on the response (``JobResultPayload`` allows ``extra``).
        if isinstance(result, LayerRun):
            dumped = result.model_dump(mode="json")
            return {
                "run_id": dumped.get("layer_run_id"),
                "run_type": "layer",
                "status": str(dumped.get("status", "")),
                "started_at": dumped.get("started_at"),
                "finished_at": dumped.get("finished_at"),
                "layer_run": dumped,
            }
        if isinstance(result, list) and result and all(
            isinstance(r, LayerRun) for r in result
        ):
            dumped = [r.model_dump(mode="json") for r in result]
            return {"target_count": len(dumped), "layer_runs": dumped}
        return _collect_payload_many(result)

    @app.get(
        f"{prefix}/jobs/{{job_id}}/stream",
        tags=["jobs"],
        response_class=StreamingResponse,
    )
    async def stream_job(job_id: str):
        """Server-Sent Events stream of a job's live state.

        Push-based alternative to polling ``GET /jobs/{job_id}``: emits one
        ``data:`` event per state/progress change (plus a keep-alive comment
        every 15s) and closes once the job reaches a terminal state. The
        client connects with ``EventSource`` and reconnects automatically.
        """
        manager = services["jobs"]
        if manager.get(job_id) is None:
            # Not live (terminal job after restart, or orphaned row). Serve
            # the persisted snapshot as ONE event and close gracefully: a
            # bare 404 here makes EventSource reconnect forever. Truly
            # unknown ids stay 404.
            persisted, _reconciled = _persisted_job_row(
                repos, manager, job_id
            )
            if persisted is None:
                raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
            payload = _job_payload(_job_from_persisted(persisted))

            async def snapshot_stream():
                yield f"data: {json.dumps(payload)}\n\n"

            return StreamingResponse(
                snapshot_stream(),
                media_type="text/event-stream",
                headers={
                    "Cache-Control": "no-cache",
                    "X-Accel-Buffering": "no",
                },
            )

        loop = asyncio.get_running_loop()
        queue = manager.subscribe(job_id, loop)

        async def event_stream():
            try:
                terminal = {"succeeded", "failed", "cancelled"}
                while True:
                    try:
                        snapshot = await asyncio.wait_for(
                            queue.get(), timeout=15.0
                        )
                    except asyncio.TimeoutError:
                        yield ": keep-alive\n\n"
                        continue
                    payload = JobPayload.model_validate(snapshot).model_dump(
                        mode="json"
                    )
                    yield f"data: {json.dumps(payload)}\n\n"
                    if snapshot["status"] in terminal:
                        return
            finally:
                manager.unsubscribe(job_id, queue)

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            },
        )

    # ------------------------------------------------------------------
    # Runs (provenance)
    # ------------------------------------------------------------------
    @app.get(f"{prefix}/runs", tags=["runs"], response_model=Paginated[RunPayload])
    def list_runs(
        run_type: RunType | None = None,
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
        sort_dir: str = Query(
            "desc",
            pattern="^(asc|desc)$",
            description="Sort direction by started_at. 'desc' (default) shows the newest runs first.",
        ),
    ):
        runs = repos.runs.list_runs(run_type=run_type)
        return _paginate(
            runs,
            cursor=cursor,
            page_size=page_size,
            key=_run_key,
            reverse=(sort_dir == "desc"),
        )

    @app.get(f"{prefix}/runs/{{run_id}}", tags=["runs"], response_model=RunPayload)
    def get_run(run_id: str):
        run = repos.runs.get_run(run_id)
        if run is None:
            raise HTTPException(status_code=404, detail=f"Run {run_id} not found")
        return _run_payload(run)

    @app.get(
        f"{prefix}/runs/{{run_id}}/sub-runs",
        tags=["runs"],
        response_model=Paginated[RunPayload],
    )
    def list_sub_runs(
        run_id: str,
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        """List the runs registered as children of this run (lineage).

        A bulk/expansion scrape creates one ``RECOMMENDATION`` run per source
        video; each is registered with ``parent_run_id`` pointing at the run
        that triggered it, so the trigger run's sub-runs are always visible.
        """
        run = repos.runs.get_run(run_id)
        if run is None:
            raise HTTPException(status_code=404, detail=f"Run {run_id} not found")
        sub_runs = repos.runs.list_sub_runs(run_id)
        return _paginate(sub_runs, cursor=cursor, page_size=page_size, key=_run_key)

    @app.patch(f"{prefix}/runs/{{run_id}}", tags=["runs"], response_model=RunPayload)
    def update_run(run_id: str, body: UpdateRunRequest):
        run = repos.runs.get_run(run_id)
        if run is None:
            raise HTTPException(status_code=404, detail=f"Run {run_id} not found")
        updates: dict[str, Any] = {}
        if body.name is not None:
            updates["name"] = body.name
        if body.tags is not None:
            updates["tags"] = [t.strip() for t in body.tags if t.strip()]
        if updates:
            run = run.model_copy(update=updates)
        repos.runs.update_run(run)
        return _run_payload(run)

    @app.patch(
        f"{prefix}/jobs/{{job_id}}/tags",
        tags=["jobs"],
    )
    def set_job_tags(job_id: str, body: JobTagsRequest):
        """Set/replace the researcher tags on a job (before/during/after)."""
        tags = [t.strip() for t in body.tags if t.strip()]
        manager: JobManager = services["jobs"]
        live = manager.get(job_id)
        if live is not None:
            live.tags = list(tags)
            manager.persist_job(live)
        else:
            persisted = _persisted_job_row(repos, services["jobs"], job_id)[0]
            if persisted is None:
                raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
            persisted.tags = list(tags)
            repos.jobs.save_job(persisted)
        return {"job_id": job_id, "tags": tags}

    @app.get(
        f"{prefix}/runs/{{run_id}}/errors",
        tags=["runs"],
        response_model=list[CollectionErrorPayload],
    )
    def run_errors(run_id: str):
        return [e.model_dump() for e in repos.runs.list_errors(run_id)]

    @app.get(
        f"{prefix}/runs/{{run_id}}/videos",
        tags=["runs"],
        response_model=RunVideosPayload,
    )
    def run_videos(
        run_id: str,
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        """List videos collected by a run.

        For a recommendation run the collected videos are the videos its edges
        observed (targets + sources, deduplicated) -- this matches the run's
        ``entities_succeeded`` edge count, so "scraped 10" shows 10 videos.
        For channel/video runs the videos are those first discovered in the
        run (``first_observed_run_id``).
        """
        run = repos.runs.get_run(run_id)
        if run is None:
            raise HTTPException(status_code=404, detail=f"Run {run_id} not found")
        if run.run_type == RunType.RECOMMENDATION:
            video_ids: set[str] = set()
            for edge in repos.recommendations.list_recommendation_edges(
                run_id=run_id
            ):
                video_ids.add(edge.source_video_id)
                video_ids.add(edge.recommended_video_id)
            videos = []
            for video_id in video_ids:
                video = repos.videos.get_video(video_id)
                if video is not None:
                    videos.append(video)
        else:
            videos = repos.videos.list_videos_by_run(run_id)
        paginated = _paginate(videos, cursor=cursor, page_size=page_size, key=_video_key)
        return RunVideosPayload(
            run_id=run_id,
            items=paginated.items,
            next_cursor=paginated.next_cursor,
            has_more=paginated.has_more,
            total=paginated.total,
        )

    # ------------------------------------------------------------------
    # Corpus / channel
    # ------------------------------------------------------------------
    @app.get(
        f"{prefix}/channels",
        tags=["corpus"],
        response_model=Paginated[ChannelPayload],
    )
    def list_channels(
        q: str | None = Query(None, description="Case-insensitive text search over title/handle/description"),
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        """All channels (global picker for comparison/explorer views)."""
        channels = repos.channels.list_channels()
        if q:
            needle = q.lower()
            channels = [
                channel
                for channel in channels
                if needle in (channel.title or "").lower()
                or needle in (channel.handle or "").lower()
                or needle in (channel.description or "").lower()
            ]
        return _paginate(channels, cursor=cursor, page_size=page_size, key=_channel_key)

    @app.get(
        f"{prefix}/channels/{{channel_id}}/overview",
        tags=["corpus"],
        response_model=ChannelOverviewPayload,
    )
    def channel_overview(channel_id: str):
        overview = services["analytics"].channel_overview(channel_id)
        return {
            "channel_id": overview.channel_id,
            "observed_at": overview.observed_at,
            "subscribers": _value_payload(overview.subscriber_count),
            "videos": _value_payload(overview.video_count),
            "views": _value_payload(overview.view_count),
        }

    @app.get(
        f"{prefix}/channels/{{channel_id}}/videos",
        tags=["corpus"],
        response_model=Paginated[VideoPayload],
    )
    def channel_videos(
        channel_id: str,
        date_from: date | None = Query(None),
        date_to: date | None = Query(None),
        video_type: str | None = Query(None),
        duration_min: int | None = Query(None),
        duration_max: int | None = Query(None),
        views_min: int | None = Query(None),
        views_max: int | None = Query(None),
        upload_hour: int | None = Query(None),
        upload_weekday: int | None = Query(None),
        keywords: list[str] | None = Query(None),
        tags: list[str] | None = Query(None),
        category: str | None = Query(None),
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        # NOTE: the VideoFilter fields are declared as explicit query params
        # (instead of FastAPI's model-as-query-param) because FastAPI cannot
        # mix a query parameter model with the pagination params below.
        filter = VideoFilter(
            date_from=date_from,
            date_to=date_to,
            video_type=video_type,
            duration_min=duration_min,
            duration_max=duration_max,
            views_min=views_min,
            views_max=views_max,
            upload_hour=upload_hour,
            upload_weekday=upload_weekday,
            keywords=keywords or [],
            tags=tags or [],
            category=category,
        )
        videos = services["query"].filter_videos(channel_id, filter)
        return _paginate(videos, cursor=cursor, page_size=page_size, key=_video_key)

    @app.get(
        f"{prefix}/channels/{{channel_id}}/videos/count",
        tags=["corpus"],
        response_model=ChannelCountPayload,
    )
    def channel_video_count(channel_id: str):
        return {
            "channel_id": channel_id,
            "count": len(repos.videos.list_videos(channel_id)),
        }

    @app.get(f"{prefix}/videos", tags=["corpus"], response_model=Paginated[VideoPayload])
    def list_videos(
        q: str | None = Query(None, description="Case-insensitive text search over title/description"),
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        """All videos (global picker for comparison/explorer views)."""
        videos = repos.videos.list_videos()
        if q:
            needle = q.lower()
            videos = [
                video
                for video in videos
                if needle in (video.title or "").lower()
                or needle in (video.description or "").lower()
            ]
        return _paginate(videos, cursor=cursor, page_size=page_size, key=_video_key)

    @app.get(f"{prefix}/videos/{{video_id}}", tags=["corpus"], response_model=VideoPayload)
    def get_video(video_id: str):
        video = repos.videos.get_video(video_id)
        if video is None:
            raise HTTPException(status_code=404, detail=f"Video {video_id} not found")
        return video.model_dump()

    @app.get(
        f"{prefix}/videos/{{video_id}}/observations",
        tags=["corpus"],
        response_model=Paginated[VideoObservationPayload],
    )
    def video_observations(
        video_id: str,
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        observations = repos.videos.list_video_observations(video_id)
        return _paginate(
            observations, cursor=cursor, page_size=page_size, key=_obs_key
        )

    @app.get(
        f"{prefix}/videos/{{video_id}}/raw",
        tags=["corpus"],
        response_model=RawVideoPayload,
    )
    def video_raw(video_id: str):
        video = repos.videos.get_video(video_id)
        if video is None:
            raise HTTPException(status_code=404, detail=f"Video {video_id} not found")
        return {"video_id": video_id, "raw_json": video.raw_json}

    @app.get(
        f"{prefix}/channels/{{channel_id}}/videos/top",
        tags=["corpus"],
        response_model=TopVideosPayload,
    )
    def channel_top_videos(
        channel_id: str,
        metric: str = Query("views"),
        n: int = Query(settings.analytics.top_n, ge=1, le=500),
    ):
        """Top videos by the latest observed engagement metric.

        Videos whose metric is MISSING are kept and annotated with
        ``availability: "missing"`` (ranked last) rather than dropped, so the
        channel-level leaderboard never silently loses videos.
        """
        field_map = {
            "views": "view_count",
            "likes": "like_count",
            "comments": "comment_count",
        }
        metric = metric.lower()
        field = field_map.get(metric, metric)
        rows: list[dict[str, Any]] = []
        for video in repos.videos.list_videos(channel_id):
            latest = repos.videos.get_latest_video_observation(video.video_id)
            value = getattr(latest, field, None) if latest is not None else None
            rows.append(
                {
                    "video_id": video.video_id,
                    "title": video.title,
                    metric: value,
                    "observed_at": latest.observed_at if latest else None,
                    "availability": "available" if value is not None else "missing",
                }
            )
        rows.sort(
            key=lambda r: (r[metric] is None, r[metric] if r[metric] is not None else 0),
            reverse=True,
        )
        return {"channel_id": channel_id, "metric": metric, "top": rows[:n]}

    # ------------------------------------------------------------------
    # Sampling
    # ------------------------------------------------------------------
    @app.post(
        f"{prefix}/channels/{{channel_id}}/videos/sample",
        tags=["sampling"],
        response_model=SamplingResultPayload,
    )
    def sample_videos(channel_id: str, spec: SamplingSpec):
        return _sampling_payload(services["sampling"].sample_videos(channel_id, spec))

    @app.post(
        f"{prefix}/videos/{{video_id}}/comments/sample",
        tags=["sampling"],
        response_model=SamplingResultPayload,
    )
    def sample_comments(video_id: str, spec: SamplingSpec):
        return _sampling_payload(services["sampling"].sample_comments(video_id, spec))

    @app.post(
        f"{prefix}/sampling/advanced",
        tags=["sampling"],
        response_model=SamplingResultPayload,
    )
    def sample_advanced(spec: AdvancedSamplingSpec):
        """Advanced sampling with complex filter combinations across channels, videos, and users.

        Supports researcher scenarios:
        - Sample/population of specific user comments across all videos and channels
        - Sample/population within specific channel(s)
        - Sample/population of specific users with their IDs
        - Sample/population of non-specified users across one channel but among specified videos
        - Video filters within same channel (date range, type, duration, views, etc.)
        - Multiple channels among specific period
        """
        return _sampling_payload(services["sampling"].sample_advanced(spec))

    # ------------------------------------------------------------------
    # Video analytics
    # ------------------------------------------------------------------
    @app.get(
        f"{prefix}/videos/{{video_id}}/engagement",
        tags=["analytics"],
        response_model=VideoEngagementPayload,
    )
    def video_engagement(video_id: str):
        eng = services["analytics"].video_engagement(video_id)
        return {
            "video_id": eng.video_id,
            "observed_at": eng.observed_at,
            "views": _value_payload(eng.views),
            "likes": _value_payload(eng.likes),
            "comments": _value_payload(eng.comments),
            "engagement_rate": _value_payload(eng.engagement_rate),
            "like_rate": _value_payload(eng.like_rate),
            "comment_rate": _value_payload(eng.comment_rate),
        }

    @app.get(
        f"{prefix}/videos/{{video_id}}/comments/percentiles",
        tags=["analytics"],
        response_model=PercentilesPayload,
    )
    def comment_percentiles(video_id: str):
        result = services["analytics"].comment_like_percentiles(video_id)
        return {
            "video_id": result.video_id,
            "availability": result.availability.value,
            "observed_like_counts": result.like_counts,
            "bands": result.bands,
        }

    @app.get(
        f"{prefix}/videos/{{video_id}}/comments/velocity",
        tags=["analytics"],
        response_model=list[VelocityPoint],
    )
    def comment_velocity(video_id: str, bucket: str = "day"):
        return services["analytics"].comment_velocity(video_id, bucket=bucket)

    @app.get(
        f"{prefix}/videos/{{video_id}}/comments",
        tags=["analytics"],
        response_model=Paginated[CommentPayload],
    )
    def video_comments(
        video_id: str,
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        comments = repos.comments.list_comments(video_id)
        page = _paginate(comments, cursor=cursor, page_size=page_size, key=_comment_key)
        # Enrich with latest observations (like_count, reply_count, is_removed)
        comment_ids = [c["comment_id"] for c in page.items]
        latest_obs = repos.comments.get_latest_comment_observations(comment_ids)
        enriched_items = []
        for c in page.items:
            obs = latest_obs.get(c["comment_id"])
            if obs:
                c["like_count"] = obs.like_count
                c["reply_count"] = obs.reply_count
                c["is_removed"] = obs.is_removed
            enriched_items.append(c)
        return Paginated(
            items=enriched_items,
            next_cursor=page.next_cursor,
            has_more=page.has_more,
            total=page.total,
        )

    @app.get(
        f"{prefix}/videos/{{video_id}}/comments/threads",
        tags=["analytics"],
        response_model=Paginated[ThreadPayload],
    )
    def video_comment_threads(
        video_id: str,
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        """Root comments with their replies, preserving parent-child order."""
        pairs = [
            (root, repos.comments.list_replies(root.comment_id))
            for root in repos.comments.list_root_comments(video_id)
        ]
        full = sorted(pairs, key=lambda pair: pair[0].comment_id)
        page = page_sorted(
            full,
            cursor=cursor,
            page_size=page_size,
            key_func=lambda pair: (pair[0].comment_id,),
            total=len(full),
        )

        # Collect all comment IDs on this page to fetch latest observations in one pass
        all_comment_ids = []
        for root, replies in page.items:
            all_comment_ids.append(root.comment_id)
            all_comment_ids.extend(r.comment_id for r in replies)
        latest_obs = repos.comments.get_latest_comment_observations(all_comment_ids)

        def enrich(comment):
            obs = latest_obs.get(comment.comment_id)
            payload = comment.model_dump()
            if obs:
                payload["like_count"] = obs.like_count
                payload["reply_count"] = obs.reply_count
                payload["is_removed"] = obs.is_removed
            return payload

        return Paginated(
            items=[
                {
                    "comment": enrich(root),
                    "replies": [enrich(r) for r in replies],
                }
                for root, replies in page.items
            ],
            next_cursor=page.next_cursor,
            has_more=page.has_more,
            total=page.total,
        )

    # ------------------------------------------------------------------
    # Comment statistics
    # ------------------------------------------------------------------
    @app.get(
        f"{prefix}/videos/{{video_id}}/comments/stats",
        tags=["analytics"],
        response_model=CommentStatsPayload,
    )
    def video_comment_stats(video_id: str):
        """Comment statistics for a video:
        - max_replies: maximum reply_count from comment observations
        - max_unique_repliers: for the comment with max replies, count distinct author_ids of its replies
        - total_replies: sum of all reply_counts
        - total_unique_repliers: count of distinct author_ids across all replies
        """
        # Get all comments for the video
        comments = repos.comments.list_comments(video_id)
        if not comments:
            return CommentStatsPayload(
                video_id=video_id,
                max_replies=0,
                max_unique_repliers=0,
                total_replies=0,
                total_unique_repliers=0,
            )

        # Get latest observations for all comments
        comment_ids = [c.comment_id for c in comments]
        latest_obs = repos.comments.get_latest_comment_observations(comment_ids)

        # Calculate total replies and find comment with max replies
        total_replies = 0
        max_replies = 0
        max_replies_comment_id = None
        for comment in comments:
            obs = latest_obs.get(comment.comment_id)
            reply_count = obs.reply_count if obs and obs.reply_count is not None else 0
            total_replies += reply_count
            if reply_count > max_replies:
                max_replies = reply_count
                max_replies_comment_id = comment.comment_id

        # Calculate total unique repliers across all replies
        # Get all replies for all comments in one pass
        all_replies = repos.comments.list_replies_by_ids(comment_ids)
        all_replier_ids: set[str] = set()
        for reply_list in all_replies.values():
            for reply in reply_list:
                if reply.author_id:
                    all_replier_ids.add(reply.author_id)
                elif reply.author_name:
                    all_replier_ids.add(reply.author_name)
        total_unique_repliers = len(all_replier_ids)

        # Calculate max_unique_repliers for the comment with max replies
        max_unique_repliers = 0
        if max_replies_comment_id:
            max_comment_replies = all_replies.get(max_replies_comment_id, [])
            max_replier_ids: set[str] = set()
            for reply in max_comment_replies:
                if reply.author_id:
                    max_replier_ids.add(reply.author_id)
                elif reply.author_name:
                    max_replier_ids.add(reply.author_name)
            max_unique_repliers = len(max_replier_ids)

        return CommentStatsPayload(
            video_id=video_id,
            max_replies=max_replies,
            max_unique_repliers=max_unique_repliers,
            total_replies=total_replies,
            total_unique_repliers=total_unique_repliers,
        )

    # ------------------------------------------------------------------
    # System folders
    # ------------------------------------------------------------------
    @app.get(
        f"{prefix}/system/folders",
        tags=["system"],
        response_model=SystemFoldersPayload,
    )
    def system_folders():
        """Return data folder paths from RepositorySettings."""
        repo_settings = settings.repository
        return SystemFoldersPayload(
            workbook_path=str(repo_settings.workbook_path),
            transcripts_dir=str(repo_settings.transcripts_dir),
            datasets_dir=str(Path(repo_settings.data_dir) / "datasets"),
            samples_dir=str(Path(repo_settings.data_dir) / "samples"),
            data_dir=repo_settings.data_dir,
        )

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    @app.post(
        f"{prefix}/export",
        tags=["export"],
        response_model=ExportResponse,
    )
    def export_data(body: ExportRequest):
        """Export selected data to Excel file.

        Two modes:
        - ``project_id`` set: export *everything the project collected* as a
          multi-sheet workbook (Videos, Comments, Channels, Recommendations, Runs).
        - otherwise: export a single entity type (optionally filtered by ``ids``/
          ``columns``) to one sheet.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from io import BytesIO

        if body.project_id:
            from SocialScienceResearch.services.export_service import (
                export_project_to_workbook,
            )

            filename, content = export_project_to_workbook(repos, body.project_id)
            return StreamingResponse(
                BytesIO(content),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"
                },
            )

        # Validate entity_type
        valid_types = {"video", "comment", "channel", "run", "sample", "dataset"}
        if body.entity_type not in valid_types:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid entity_type. Must be one of: {', '.join(sorted(valid_types))}",
            )

        # Get data based on entity_type
        data: list[dict[str, Any]] = []
        if body.entity_type == "video":
            if body.ids:
                for vid in body.ids:
                    v = repos.videos.get_video(vid)
                    if v:
                        data.append(v.model_dump())
            else:
                data = [v.model_dump() for v in repos.videos.list_videos()]
        elif body.entity_type == "comment":
            if body.ids:
                for cid in body.ids:
                    c = repos.comments.get_comment(cid)
                    if c:
                        data.append(c.model_dump())
            else:
                data = [c.model_dump() for c in repos.comments.list_comments()]
        elif body.entity_type == "channel":
            if body.ids:
                for chid in body.ids:
                    ch = repos.channels.get_channel(chid)
                    if ch:
                        data.append(ch.model_dump())
            else:
                data = [ch.model_dump() for ch in repos.channels.list_channels()]
        elif body.entity_type == "run":
            if body.ids:
                for rid in body.ids:
                    r = repos.runs.get_run(rid)
                    if r:
                        data.append(r.model_dump())
            else:
                data = [r.model_dump() for r in repos.runs.list_runs()]
        elif body.entity_type == "sample":
            if body.ids:
                for sid in body.ids:
                    s = repos.samples.get(sid)
                    if s:
                        data.append(s.model_dump())
            else:
                data = [s.model_dump() for s in repos.samples.list()]
        elif body.entity_type == "dataset":
            if body.ids:
                for did in body.ids:
                    d = repos.datasets.get_dataset(did)
                    if d:
                        data.append(d.model_dump())
            else:
                data = [d.model_dump() for d in repos.datasets.list_datasets()]

        # Filter columns if specified
        if body.columns and data:
            filtered_data = []
            for row in data:
                filtered_data.append({col: row.get(col) for col in body.columns if col in row})
            data = filtered_data

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = body.entity_type.capitalize()

        if data:
            # Write headers
            headers = list(data[0].keys())
            header_font = Font(bold=True)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font

            # Write data rows
            for row_idx, row in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header)
                    # Convert complex types to string
                    if isinstance(value, (dict, list)):
                        value = str(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        filename = body.filename or f"{body.entity_type}_export.xlsx"

        # Return as file response
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # ------------------------------------------------------------------
    # Recommendation network
    # ------------------------------------------------------------------
    @app.get(
        f"{prefix}/videos/{{video_id}}/recommendations",
        tags=["network"],
        response_model=Paginated[RecommendationPayload],
    )
    def video_recommendations(
        video_id: str,
        cursor: str | None = Query(None, description="Opaque cursor from the previous page"),
        page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    ):
        edges = repos.recommendations.list_recommendations_for_source(video_id)
        # Enrich with run_type from the collection run
        enriched_edges = []
        for edge in edges:
            run = repos.runs.get_run(edge.collection_run_id)
            run_type = run.run_type.value if run else None
            edge_dict = edge.model_dump()
            edge_dict["run_type"] = run_type
            enriched_edges.append(edge_dict)
        return _paginate(
            enriched_edges, cursor=cursor, page_size=page_size, key=_recommendation_edge_key
        )

    @app.get(
        f"{prefix}/network/recommendations/summary",
        tags=["network"],
        response_model=NetworkSummaryPayload,
    )
    def network_summary(
        run_id: str | None = None,
        top_n: int = Query(settings.analytics.top_n, ge=1, le=500),
    ):
        return services["network"].summary(run_id=run_id, top_n=top_n).__dict__

    @app.get(
        f"{prefix}/network/recommendations/{{video_id}}",
        tags=["network"],
        response_model=VideoNetworkContextPayload,
    )
    def network_video_context(
        video_id: str,
        run_id: str | None = None,
        run_ids: str | None = None,
    ):
        run_id_list = [r for r in run_ids.split(",") if r] if run_ids else None
        return (
            services["network"]
            .video_context(video_id, run_id=run_id, run_ids=run_id_list)
            .__dict__
        )

    # ------------------------------------------------------------------
    # Quality / coverage
    # ------------------------------------------------------------------
    @app.get(
        f"{prefix}/coverage",
        tags=["quality"],
        response_model=CoverageReport,
    )
    def coverage():
        return services["quality"].coverage()

    @app.get(
        f"{prefix}/dataset/summary",
        tags=["quality"],
        response_model=DatasetSummaryPayload,
    )
    def dataset_summary():
        return services["quality"].dataset_summary()

    # ------------------------------------------------------------------
    # Research queries (B1): variable catalogue, operators and the funnel
    # ------------------------------------------------------------------
    @app.get(
        f"{prefix}/research/variables",
        tags=["research"],
        response_model=list[VariableMetaPayload],
    )
    def research_variables(entity: str | None = Query(None)):
        """Registered research variables for an entity (all when omitted)."""
        if entity is None:
            return [v.model_dump() for v in VariableRegistry.all_variables()]
        return [v.model_dump() for v in VariableRegistry.get_variables(entity)]

    @app.get(
        f"{prefix}/research/operators",
        tags=["research"],
        response_model=list[OperatorInfoPayload],
    )
    def research_operators():
        """Operators understood by the research-query evaluator."""
        return [
            {"name": operator.value, "description": description}
            for operator, description in OPERATOR_DESCRIPTIONS.items()
        ]

    @app.post(
        f"{prefix}/research/query/preview",
        tags=["research"],
        response_model=QueryPreviewResponse,
    )
    def research_query_preview(body: ResearchQueryRequest):
        """Evaluate a research query and report the ordered funnel stages.

        ``stages`` flatten the condition tree: each stage's ``cumulative`` is
        the count matching the conditions-so-far (AND-ed prefix) and
        ``matched`` is the incremental drop. OR/NOT groups appear once.
        """
        rows = services["query"].resolve_latest_rows(
            body.entity, context=body.query_context
        )
        preview = preview_query(body.entity, body.root, rows)
        return {
            "total": preview.total,
            "stages": [stage.model_dump() for stage in preview.stages],
            "population_size": preview.population_size,
            "n": preview.n,
        }

    @app.post(
        f"{prefix}/research/query/resolve",
        tags=["research"],
        response_model=QueryResolveResponse,
    )
    def research_query_resolve(body: ResearchQueryRequest):
        """Count-only resolution of a research query (no rows returned)."""
        rows = services["query"].resolve_latest_rows(
            body.entity, context=body.query_context
        )
        matched = evaluate_query(body.entity, body.root, rows)
        return {"total": len(matched), "population_size": len(rows)}

    return app


def _collection_payload(result) -> dict[str, Any]:
    return {
        "run_id": result.run_id,
        "run_type": result.run_type.value,
        "status": result.status.value,
        "target_url": result.target_url,
        "target_id": result.target_id,
        "entities_discovered": result.entities_discovered,
        "entities_created": result.entities_created,
        "entities_existing": result.entities_existing,
        "entities_failed": result.entities_failed,
        "comments_collected": result.comments_collected,
        "errors": [e.model_dump() for e in result.errors],
        "skipped": result.skipped,
        "started_at": result.started_at,
        "finished_at": result.finished_at,
        "dataset_id": result.dataset_id,
    }


def _collect_payload_many(results) -> dict[str, Any]:
    """Serialize one or many collection results into a uniform payload."""
    if isinstance(results, list):
        return {
            "target_count": len(results),
            "results": [_collection_payload(r) for r in results],
        }
    return _collection_payload(results)


def _run_payload(run) -> dict[str, Any]:
    return run.model_dump()


def _value_payload(value) -> dict[str, Any]:
    return {
        "value": value.value,
        "availability": value.availability.value,
    }


def _sampling_payload(result) -> dict[str, Any]:
    return {
        "strategy": result.strategy.value,
        "entity_type": result.entity_type,
        "population_size": result.population_size,
        "sample_size": result.sample_size,
        "entity_ids": result.entity_ids,
        "criteria_json": result.criteria_json,
        "seed": result.seed,
        "missing_metric_count": result.missing_metric_count,
    }


#: Module-level app so ``from SocialScienceResearch.api.app import app`` works
#: (used by the S0 import gate). Constructed with default settings; the
#: workbook store is opened in-memory and only written on an explicit save.
app = create_app()
