"""B7 tests: DatasetService / ProjectService + the dataset & project endpoints.

Covers project CRUD + ``config_hash`` stability, dataset construction (plain
corpus snapshot and from a project's research query + variable selection),
member chunking with 50k ids (ADR-0001), quality missing-share math, CSV/JSON
export markers, the raw-payload sidecar, and the FastAPI endpoints via
TestClient (pagination envelope, error envelope, streaming export).
"""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from SocialScienceResearch.api import create_app
from SocialScienceResearch.config.settings import (
    ApiSettings,
    CollectionSettings,
    RepositorySettings,
    ScraperSettings,
    SocialScienceSettings,
)
from SocialScienceResearch.domain.dataset_models import (
    CreateProjectRequest,
    Dataset,
    Project,
    UpdateProjectRequest,
)
from SocialScienceResearch.domain.models import (
    Channel,
    ChannelObservation,
    CollectionRun,
    Video,
    VideoObservation,
)
from SocialScienceResearch.domain.query import (
    Operator,
    QueryCondition,
    QueryGroup,
    ResearchQueryRequest,
)
from SocialScienceResearch.persistence.dataset_repository import DatasetRepository
from SocialScienceResearch.persistence.excel_repository import build_excel_repositories
from SocialScienceResearch.services.dataset_service import DatasetService
from SocialScienceResearch.services.project_service import ProjectService
from SocialScienceResearch.utils.idgen import utcnow

PREFIX = "/api/v1/social-science"
CHANNEL_ID = "UCb7000000000000000000000"


def _repo_settings(tmp_path) -> RepositorySettings:
    return RepositorySettings(data_dir=str(tmp_path), dataset_name="b7")


def _build_settings(repo_settings: RepositorySettings) -> SocialScienceSettings:
    return SocialScienceSettings(
        repository=repo_settings,
        scraper=ScraperSettings(retries=1, retry_backoff=0.0, request_delay_seconds=0),
        collection=CollectionSettings(collect_comments=False),
        api=ApiSettings(prefix=PREFIX),
    )


def _seed_corpus(repos) -> None:
    """One channel + 5 observed videos + 1 video without an observation."""
    repos.runs.create_run(
        CollectionRun(
            run_id="run_b7",
            run_type="channel",
            target_url="https://www.youtube.com/@b7",
            target_channel_id=CHANNEL_ID,
            started_at=utcnow(),
            status="success",
        )
    )
    repos.channels.upsert_channel(
        Channel(
            channel_id=CHANNEL_ID,
            url=f"https://www.youtube.com/channel/{CHANNEL_ID}",
            title="B7 Channel",
            first_observed_run_id="run_b7",
        )
    )
    repos.channels.save_channel_observation(
        ChannelObservation(
            observation_id="obs_b7_ch",
            collection_run_id="run_b7",
            channel_id=CHANNEL_ID,
            observed_at=utcnow(),
            subscriber_count=500,
            video_count=6,
            view_count=1_000_000,
        )
    )
    for i in range(5):
        vid = f"v0{i}"
        repos.videos.upsert_video(
            Video(
                video_id=vid,
                url=f"https://www.youtube.com/watch?v={vid}",
                channel_id=CHANNEL_ID,
                title=f"B7 video {i}",
                description=f"research sample {i}",
                duration=100 + i,
                upload_timestamp=utcnow(),
                tags=["research"],
                is_short=i % 2 == 0,
                raw_json={"title": f"raw {vid}"},
                first_observed_run_id="run_b7",
            )
        )
        repos.videos.save_video_observation(
            VideoObservation(
                observation_id=f"obs_v_{vid}",
                collection_run_id="run_b7",
                video_id=vid,
                observed_at=utcnow(),
                view_count=100 + i * 100,  # 100, 200, 300, 400, 500
                like_count=10 + i,
                comment_count=2 + i,
            )
        )
    # One video with no observation: view-derived columns stay None.
    repos.videos.upsert_video(
        Video(
            video_id="v_noobs",
            url="https://www.youtube.com/watch?v=v_noobs",
            channel_id=CHANNEL_ID,
            title="No observations",
            duration=99,
            first_observed_run_id="run_b7",
        )
    )


@pytest.fixture
def service_env(tmp_path):
    """Open Excel repositories + settings for direct service-level tests."""
    repo_settings = _repo_settings(tmp_path)
    repos = build_excel_repositories(repo_settings)
    _seed_corpus(repos)
    return repos, _build_settings(repo_settings)


@pytest.fixture
def client(tmp_path):
    repo_settings = _repo_settings(tmp_path)
    repos = build_excel_repositories(repo_settings)
    _seed_corpus(repos)
    repos.store.close()
    app = create_app(settings=_build_settings(repo_settings))
    with TestClient(app) as test_client:
        yield test_client


def _make_project(
    projects: ProjectService, project_id: str, **overrides
) -> Project:
    request = CreateProjectRequest(
        name=overrides.get("name", f"Project {project_id}"),
        description=None,
        targets=[{"kind": "channel", "url": "https://www.youtube.com/@b7"}],
        collection_spec={"collect_comments": True},
        sampling_specs=[{"strategy": "random", "size": 10}],
        research_query=ResearchQueryRequest(
            entity="video",
            root=QueryGroup(
                operator="AND",
                conditions=[
                    QueryCondition(
                        variable="view_count", operator=Operator.GT, value=200
                    )
                ],
            ),
        ).model_dump(),
        variable_selection=sorted(overrides.get("variable_selection", ["video_id"])),
        notes=None,
    )
    now = utcnow()
    return projects.create(
        Project(
            project_id=project_id,
            name=request.name,
            description=request.description,
            targets=[dict(t) for t in request.targets],
            collection_spec=dict(request.collection_spec),
            sampling_specs=[dict(s) for s in request.sampling_specs],
            research_query=dict(request.research_query),
            variable_selection=list(request.variable_selection),
            notes=request.notes,
            config_hash="",
            created_at=now,
            updated_at=now,
        )
    )


# ----------------------------------------------------------------------
# Dataset construction
# ----------------------------------------------------------------------
def test_create_dataset_snapshot(service_env) -> None:
    repos, settings = service_env
    service = DatasetService(repos, settings)
    dataset = service.create_dataset(
        "all videos", "full population", entity_type="video"
    )
    assert dataset.entity_type == "video"
    assert dataset.member_count == 6
    assert dataset.overflow is False
    assert dataset.created_by_run_id == "run_b7"
    assert dataset.source_projection["id_field"] == "video_id"
    assert dataset.source_projection["project_id"] is None

    members = service.members(dataset.dataset_id)
    assert [m["video_id"] for m in members] == [
        "v00", "v01", "v02", "v03", "v04", "v_noobs",
    ]
    fetched = service.get_dataset(dataset.dataset_id)
    assert fetched == dataset
    assert service.member_count(dataset.dataset_id) == 6


def test_create_dataset_member_ids_scopes_rows(service_env) -> None:
    repos, settings = service_env
    service = DatasetService(repos, settings)
    dataset = service.create_dataset(
        "sample subset",
        entity_type="video",
        member_ids=["v01", "v03"],
    )
    assert dataset.member_count == 2
    assert dataset.source_projection["scope"]["member_ids"] == ["v01", "v03"]
    members = service.members(dataset.dataset_id)
    assert {m["video_id"] for m in members} == {"v01", "v03"}


def test_create_from_project_honors_query_and_variable_selection(service_env) -> None:
    repos, settings = service_env
    projects = ProjectService(repos)
    project = _make_project(
        projects, "proj_high", variable_selection=["video_id", "title", "view_count"]
    )

    datasets = DatasetService(repos, settings)
    dataset = datasets.create_from_project("proj_high")
    # GT 200 over view_count [100..500] + v_noobs(None) -> v02 300, v03 400, v04 500
    assert dataset.member_count == 3
    assert dataset.source_projection["project_id"] == "proj_high"
    assert dataset.source_projection["query_hash"]  # documented projection
    members = datasets.members(dataset.dataset_id)
    assert all(set(m) == {"video_id", "title", "view_count"} for m in members)
    assert all(set(m) == {"video_id", "title", "view_count"} for m in members)
    assert {m["video_id"] for m in members} == {"v02", "v03", "v04"}
    assert all(m["view_count"] > 200 for m in members)


def test_include_raw_writes_sidecar(service_env, tmp_path) -> None:
    repos, settings = service_env
    datasets = DatasetService(repos, settings)
    dataset = datasets.create_dataset(
        "raw snapshot", entity_type="video", include_raw=True
    )
    raw_path = Path(settings.repository.data_dir) / "raw" / f"{dataset.dataset_id}.json"
    assert raw_path.exists()
    payload = json.loads(raw_path.read_text(encoding="utf-8"))
    assert set(payload) == {"v00", "v01", "v02", "v03", "v04", "v_noobs"}
    assert payload["v00"]["title"] == "raw v00"


def test_unknown_entity_raises(service_env) -> None:
    repos, settings = service_env
    service = DatasetService(repos, settings)
    with pytest.raises(ValueError):
        service.create_dataset("bad", entity_type="planet")
    with pytest.raises(ValueError):
        service.get_dataset("missing")
    with pytest.raises(ValueError):
        service.delete_dataset("missing")


# ----------------------------------------------------------------------
# Member chunking (ADR-0001 ~32k cell limit)
# ----------------------------------------------------------------------
def test_member_chunking_50k_ids(tmp_path) -> None:
    repo_settings = _repo_settings(tmp_path)
    repos = build_excel_repositories(repo_settings)
    repository = DatasetRepository(repos.store)

    dataset = Dataset(
        dataset_id="dst_big",
        name="big",
        entity_type="video",
        created_at=utcnow(),
        source_projection={"id_field": "video_id", "columns": ["video_id"]},
        member_count=50000,
    )
    repository.save_dataset(dataset)

    members = [{"video_id": f"member-{i:06d}"} for i in range(50_000)]
    chunks = repository.save_members("dst_big", members)
    assert chunks > 1, "50k members must overflow a single chunk"

    stored = repository.list_members("dst_big")
    assert len(stored) == 50_000
    assert repository.dataset_member_count("dst_big") == 50_000
    assert {m["video_id"] for m in stored} == {f"member-{i:06d}" for i in range(50_000)}

    fetched = repository.get_dataset("dst_big")
    assert fetched is not None
    assert fetched.member_count == 50_000
    assert fetched.dataset_id == "dst_big"


def test_small_dataset_uses_single_chunk(tmp_path) -> None:
    repo_settings = _repo_settings(tmp_path)
    repos = build_excel_repositories(repo_settings)
    repository = DatasetRepository(repos.store)
    chunks = repository.save_members("dst_small", [{"video_id": "a"}, {"video_id": "b"}])
    assert chunks == 1
    assert repository.dataset_member_count("dst_small") == 2


# ----------------------------------------------------------------------
# Quality
# ----------------------------------------------------------------------
def test_quality_missing_share_math(service_env) -> None:
    repos, settings = service_env
    service = DatasetService(repos, settings)
    dataset = service.create_dataset("quality", entity_type="video")
    report = service.quality(dataset.dataset_id)
    assert report.dataset_id == dataset.dataset_id
    n = 6
    columns = {column.name: column for column in report.columns}

    assert columns["video_id"].present == n
    assert columns["video_id"].missing == 0
    # Only v00..v04 carry latest observations; v_noobs has none.
    assert columns["view_count"].present == 5
    assert columns["view_count"].missing == 1
    assert columns["view_count"].missing_share == pytest.approx(1 / 6, abs=1e-4)
    assert columns["like_count"].present == 5

    present_total = sum(column.present for column in report.columns)
    assert report.overall_coverage == pytest.approx(
        present_total / (n * len(report.columns)), abs=1e-4
    )
    assert report.corpus["videos"] == 6  # reused QualityService.dataset_summary()


# ----------------------------------------------------------------------
# Export
# ----------------------------------------------------------------------
def test_export_csv_and_json_markers(service_env) -> None:
    repos, settings = service_env
    service = DatasetService(repos, settings)
    dataset = service.create_dataset("exportable", entity_type="video")

    filename, content, media_type = service.export(dataset.dataset_id, "csv")
    assert filename.endswith(".csv")
    assert media_type == "text/csv"
    lines = content.splitlines()
    assert lines[0].startswith("video_id,")
    assert any("B7 video 1" in line for line in lines)
    assert any(line.startswith("v_noobs,") for line in lines)


def test_export_project_to_workbook_multi_sheet(service_env) -> None:
    """A project's collected data exports as one multi-sheet Excel workbook."""
    from openpyxl import load_workbook

    from SocialScienceResearch.domain.dataset_models import Project, ProjectItem
    from SocialScienceResearch.domain.models import RecommendationObservation
    from SocialScienceResearch.services.export_service import (
        export_project_to_workbook,
    )
    from SocialScienceResearch.services.project_service import ProjectService

    repos, settings = service_env
    dataset = DatasetService(repos, settings).create_dataset(
        "collected videos", entity_type="video"
    )

    # A recommendation edge among the project's videos (so the Recommendations
    # sheet is populated and run provenance is captured).
    repos.recommendations.save_recommendation(
        RecommendationObservation(
            observation_id="rec_export",
            collection_run_id="run_b7",
            source_video_id="v00",
            recommended_video_id="v01",
            position=0,
            title="B7 video 1",
            channel_id=CHANNEL_ID,
        )
    )

    now = utcnow()
    project = ProjectService(repos).create(
        Project(
            project_id="proj_export",
            name="Export project",
            description=None,
            targets=[],
            collection_spec={},
            sampling_specs=[],
            research_query=None,
            variable_selection=[],
            notes=None,
            config_hash="",
            created_at=now,
            updated_at=now,
        )
    )
    repos.project_items.save_item(
        ProjectItem(
            item_id="item_export",
            project_id=project.project_id,
            name="Collected videos",
            item_type="dataset_group",
            dataset_ids=[dataset.dataset_id],
            created_at=now,
            updated_at=now,
        )
    )

    filename, content = export_project_to_workbook(repos, project.project_id)
    assert filename == "project_proj_export_export.xlsx"

    wb = load_workbook(io.BytesIO(content))
    assert "Videos" in wb.sheetnames
    assert "Runs" in wb.sheetnames
    assert "Recommendations" in wb.sheetnames

    videos = wb["Videos"]
    # 1 header + 6 seeded videos
    assert videos.max_row == 7
    video_ids = {v[0].value for v in videos.iter_rows(min_row=2, max_col=1)}
    assert {"v00", "v01", "v_noobs"} <= video_ids

    recommendations = wb["Recommendations"]
    assert recommendations.max_row >= 2  # header + the v00 -> v01 edge
    rec_pairs = {
        (r[0].value, r[1].value)
        for r in recommendations.iter_rows(min_row=2, max_col=2)
    }
    assert ("v00", "v01") in rec_pairs

    runs = wb["Runs"]
    run_ids = {r[0].value for r in runs.iter_rows(min_row=2, max_col=1)}
    assert "run_b7" in run_ids

# ----------------------------------------------------------------------
# Projects
# ----------------------------------------------------------------------
def test_project_crud_and_config_hash(service_env) -> None:
    repos, _ = service_env
    projects = ProjectService(repos)

    project = _make_project(projects, "proj_a", variable_selection=["video_id"])
    assert project.config_hash and len(project.config_hash) == 16
    # Same definition -> same hash (stable across independent creations).
    twin = _make_project(
        projects, "proj_a2", name="Project proj_a", variable_selection=["video_id"]
    )
    assert twin.config_hash == project.config_hash

    updated = projects.update_project("proj_a", UpdateProjectRequest(name="Renamed"))
    assert updated.name == "Renamed"
    assert updated.config_hash != project.config_hash
    assert updated.updated_at >= project.updated_at
    assert projects.get_project("proj_a").name == "Renamed"

    selected = projects.update_project(
        "proj_a", UpdateProjectRequest(variable_selection=["video_id", "title"])
    )
    assert selected.variable_selection == ["video_id", "title"]
    assert selected.config_hash != updated.config_hash
    assert [p.project_id for p in projects.list_projects()] == [
        "proj_a", "proj_a2",
    ]

    projects.delete_project("proj_a")
    with pytest.raises(ValueError):
        projects.get_project("proj_a")
    assert [p.project_id for p in projects.list_projects()] == ["proj_a2"]


# ----------------------------------------------------------------------
# API endpoints (TestClient)
# ----------------------------------------------------------------------
def test_api_project_and_dataset_flow(client) -> None:
    resp = client.post(
        f"{PREFIX}/projects",
        json={
            "name": "API project",
            "targets": [{"kind": "channel", "url": "https://www.youtube.com/@b7"}],
            "research_query": {
                "entity": "video",
                "root": {
                    "operator": "AND",
                    "conditions": [
                        {"variable": "view_count", "operator": "gt", "value": 200}
                    ],
                },
            },
            "variable_selection": ["video_id", "title", "view_count"],
        },
    )
    assert resp.status_code == 200, resp.text
    project_id = resp.json()["project_id"]
    assert resp.json()["config_hash"]

    resp = client.get(f"{PREFIX}/projects")
    assert resp.status_code == 200
    assert set(resp.json()) == {"items", "next_cursor", "has_more", "total"}
    assert resp.json()["total"] == 1

    resp = client.get(f"{PREFIX}/projects/{project_id}")
    assert resp.status_code == 200
    assert resp.json()["name"] == "API project"

    hash_before = resp.json()["config_hash"]
    resp = client.patch(f"{PREFIX}/projects/{project_id}", json={"name": "API v2"})
    assert resp.status_code == 200
    assert resp.json()["name"] == "API v2"
    assert resp.json()["config_hash"] != hash_before

    resp = client.post(
        f"{PREFIX}/datasets",
        json={"name": "api ds", "entity_type": "video", "project_id": project_id},
    )
    assert resp.status_code == 200, resp.text
    dataset_id = resp.json()["dataset_id"]
    assert resp.json()["member_count"] == 3

    resp = client.get(f"{PREFIX}/datasets")
    assert resp.status_code == 200
    assert set(resp.json()) == {"items", "next_cursor", "has_more", "total"}

    resp = client.get(f"{PREFIX}/datasets/{dataset_id}/members")
    assert resp.status_code == 200
    body = resp.json()
    assert body["total"] == 3
    assert {item["video_id"] for item in body["items"]} == {"v02", "v03", "v04"}

    resp = client.get(f"{PREFIX}/datasets/{dataset_id}/quality")
    assert resp.status_code == 200
    assert resp.json()["dataset_id"] == dataset_id
    assert resp.json()["overall_coverage"] >= 0

    resp = client.get(
        f"{PREFIX}/datasets/{dataset_id}/export", params={"format": "csv"}
    )
    assert resp.status_code == 200
    assert "video_id,title,view_count" in resp.text
    assert "text/csv" in resp.headers["content-type"]
    assert resp.headers["content-disposition"].startswith("attachment")

    resp = client.get(
        f"{PREFIX}/datasets/{dataset_id}/export", params={"format": "json"}
    )
    assert resp.status_code == 200
    assert "application/json" in resp.headers["content-type"]
    assert json.loads(resp.text)["dataset"]["dataset_id"] == dataset_id

    resp = client.delete(f"{PREFIX}/datasets/{dataset_id}")
    assert resp.status_code == 200
    assert resp.json()["deleted"] is True
    resp = client.get(f"{PREFIX}/datasets/{dataset_id}")
    assert resp.status_code == 400
    assert resp.json()["code"] == "invalid_argument"

    # ValueError -> 400 envelope via the app's handler.
    resp = client.post(
        f"{PREFIX}/datasets",
        json={"name": "bad", "entity_type": "video", "project_id": "missing_proj"},
    )
    assert resp.status_code == 400
    assert resp.json()["code"] == "invalid_argument"

    resp = client.delete(f"{PREFIX}/projects/{project_id}")
    assert resp.status_code == 200
    resp = client.get(f"{PREFIX}/projects/{project_id}")
    assert resp.status_code == 400


def test_api_project_items_flow(client) -> None:
    """Project items CRUD works end-to-end (regression: service must not touch
    a non-existent ``repos.projects`` attribute)."""
    resp = client.post(
        f"{PREFIX}/projects",
        json={
            "name": "Items project",
            "targets": [{"kind": "channel", "url": "https://www.youtube.com/@b7"}],
            "variable_selection": ["video_id", "view_count"],
        },
    )
    assert resp.status_code == 200, resp.text
    project_id = resp.json()["project_id"]

    resp = client.get(f"{PREFIX}/projects/{project_id}/items")
    assert resp.status_code == 200
    assert resp.json() == []

    resp = client.post(
        f"{PREFIX}/projects/{project_id}/items",
        json={
            "name": "Item A",
            "item_type": "sample_group",
            "sample_ids": ["s1"],
            "dataset_ids": [],
            "tags": ["t1"],
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    item_id = body["item_id"]
    assert body["project_id"] == project_id
    assert body["name"] == "Item A"
    assert body["sample_ids"] == ["s1"]

    resp = client.post(
        f"{PREFIX}/projects/{project_id}/items",
        json={"name": "Item B", "item_type": "mixed", "sample_ids": [], "dataset_ids": ["d1"]},
    )
    assert resp.status_code == 200, resp.text

    resp = client.get(f"{PREFIX}/projects/{project_id}/items")
    assert resp.status_code == 200
    assert {item["name"] for item in resp.json()} == {"Item A", "Item B"}

    resp = client.get(f"{PREFIX}/projects/{project_id}/items/{item_id}")
    assert resp.status_code == 200
    assert resp.json()["name"] == "Item A"

    resp = client.patch(
        f"{PREFIX}/projects/{project_id}/items/{item_id}",
        json={"name": "Item A v2"},
    )
    assert resp.status_code == 200
    assert resp.json()["name"] == "Item A v2"

    resp = client.delete(f"{PREFIX}/projects/{project_id}/items/{item_id}")
    assert resp.status_code == 200
    assert resp.json()["deleted"] is True

    resp = client.get(f"{PREFIX}/projects/{project_id}/items")
    assert resp.status_code == 200
    assert {item["name"] for item in resp.json()} == {"Item B"}