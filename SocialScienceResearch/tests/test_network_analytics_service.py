"""Tests for ``NetworkAnalyticsService`` and the B6 network router.

Seeds a small deterministic recommendation network across two runs:

* ``net_r1``: a single reciprocated pair ``a <-> b``;
* ``net_r2``: ``a2->b2, a2->c2, b2->c2, c2->a2, d2->a2`` (5 edges).

The combined (``run_id=None``) graph therefore has 6 nodes and 7 edges in two
weakly-connected components.
"""

from __future__ import annotations

import json

import pytest
from fastapi.testclient import TestClient

from SocialScienceResearch.api import create_app
from SocialScienceResearch.config.settings import (
    ApiSettings,
    CollectionSettings,
    RepositorySettings,
    ScraperSettings,
    SocialScienceSettings,
)
from SocialScienceResearch.domain.enums import RecommendationStatus, RunType
from SocialScienceResearch.domain.models import (
    Channel,
    CollectionRun,
    RecommendationObservation,
    Video,
)
from SocialScienceResearch.persistence.excel_repository import build_excel_repositories
from SocialScienceResearch.services.network_analytics_service import (
    NetworkAnalyticsService,
    NetworkScope,
)
from SocialScienceResearch.utils.idgen import utcnow

PREFIX = "/api/v1/social-science"


def _seed_recommendations(repos) -> None:
    """Seed the deterministic 2-run network described in the module docstring."""
    edges = [
        ("r_obs_1", "net_r1", "a", "b", 0, "UC1", "T a->b"),
        ("r_obs_2", "net_r1", "b", "a", 1, "UC1", "T b->a"),
        ("r_obs_3", "net_r2", "a2", "b2", 0, "UC2", "T a2->b2"),
        ("r_obs_4", "net_r2", "a2", "c2", 1, "UC3", "T a2->c2"),
        ("r_obs_5", "net_r2", "b2", "c2", 0, "UC3", "T b2->c2"),
        ("r_obs_6", "net_r2", "c2", "a2", 2, "UC2", "T c2->a2"),
        ("r_obs_7", "net_r2", "d2", "a2", 3, "UC2", "T d2->a2"),
    ]
    for observation_id, run_id, source, target, position, channel_id, title in edges:
        repos.recommendations.save_recommendation(
            RecommendationObservation(
                observation_id=observation_id,
                collection_run_id=run_id,
                source_video_id=source,
                recommended_video_id=target,
                position=position,
                status=RecommendationStatus.OBSERVED,
                channel_id=channel_id,
                title=title,
            )
        )


def _seed_videos(repos) -> None:
    """Persist source video rows so source-channel metadata resolves."""
    videos = [
        ("a", "UC1"),
        ("b", "UC1"),
        ("a2", "UC2"),
        ("b2", "UC3"),
        ("c2", "UC3"),
        ("d2", "UC2"),
    ]
    for video_id, channel_id in videos:
        repos.videos.upsert_video(
            Video(
                video_id=video_id,
                url=f"https://www.youtube.com/watch?v={video_id}",
                channel_id=channel_id,
                title=f"Title {video_id}",
                first_observed_run_id="net_r1",
            )
        )
    for channel_id in ("UC1", "UC2", "UC3"):
        repos.channels.upsert_channel(
            Channel(
                channel_id=channel_id,
                url=f"https://www.youtube.com/channel/{channel_id}",
                title=f"Channel {channel_id}",
                first_observed_run_id="net_r1",
            )
        )


def _seed_runs(repos) -> None:
    """Persist run rows so the graph returns the run facet."""
    for run_id, run_type in (("net_r1", RunType.VIDEO), ("net_r2", RunType.VIDEO)):
        repos.runs.create_run(
            CollectionRun(
                run_id=run_id,
                run_type=run_type,
                target_url=f"https://www.youtube.com/watch?v={run_id}",
                started_at=utcnow(),
                status="success",
            )
        )


@pytest.fixture
def service(excel_repos) -> NetworkAnalyticsService:
    # Remove the import that's causing issues
    _seed_recommendations(excel_repos)
    return NetworkAnalyticsService(excel_repos)


@pytest.fixture
def client(tmp_path):
    repo_settings = RepositorySettings(data_dir=str(tmp_path), dataset_name="net")
    repos = build_excel_repositories(repo_settings)
    _seed_recommendations(repos)
    repos.store.close()

    settings = SocialScienceSettings(
        repository=repo_settings,
        scraper=ScraperSettings(retries=1, retry_backoff=0.0, request_delay_seconds=0),
        collection=CollectionSettings(collect_comments=False),
        api=ApiSettings(prefix=PREFIX),
    )
    app = create_app(settings)
    yield TestClient(app)


# ----------------------------------------------------------------------
# Service: metrics
# ----------------------------------------------------------------------
def test_metrics_full_graph_counts_and_bounds(service) -> None:
    metrics = service.metrics()
    assert metrics.node_count == 6
    assert metrics.edge_count == 7
    assert metrics.is_directed is True
    assert 0.0 <= metrics.density <= 1.0
    assert 0.0 <= metrics.reciprocity <= 1.0
    assert metrics.weakly_connected_components == 2
    assert metrics.largest_component_size == 4
    assert metrics.largest_component_share == pytest.approx(4 / 6)


def test_metrics_reciprocity_bidirectional_pair(service) -> None:
    metrics = service.metrics(run_id="net_r1")
    assert metrics.node_count == 2
    assert metrics.edge_count == 2
    assert metrics.reciprocity == 1.0
    assert metrics.density == 1.0


def test_run_scoped_metrics_match_graph_view(service) -> None:
    """Metrics must be computed from the same edge set the graph view renders.

    Regression guard for the bug where ``metrics()`` used the cached raw
    ``build_graph`` (which could disagree with / lag the displayed graph), so
    density / reciprocity / clustering / components / communities / HITS did not
    match the interactive graph. The metrics panel must agree with the graph.
    """
    import networkx as nx

    for run_id in (None, "net_r1", "net_r2"):
        metrics = service.metrics(run_id=run_id)
        graph = service.graph(run_id=run_id)

        assert metrics.node_count == graph.node_count == len(graph.nodes)
        assert metrics.edge_count == graph.edge_count == len(graph.edges)

        g = nx.DiGraph()
        for e in graph.edges:
            g.add_edge(e.source, e.target)
        assert metrics.weakly_connected_components == len(
            list(nx.weakly_connected_components(g))
        )


def test_metrics_degree_percentiles_on_known_distribution(service) -> None:
    metrics = service.metrics(run_id="net_r2")
    assert metrics.node_count == 4
    assert metrics.edge_count == 5

    in_deg = metrics.degree_distribution["in_degree"]
    # in-degrees across net_r2 nodes: a2=2, b2=1, c2=2, d2=0 -> [0, 1, 2, 2].
    assert in_deg.min == 0
    assert in_deg.max == 2
    assert in_deg.mean == pytest.approx(1.25)
    assert in_deg.median == pytest.approx(1.5)
    assert in_deg.p25 == pytest.approx(0.75)
    assert in_deg.p75 == pytest.approx(2.0)

    out_deg = metrics.degree_distribution["out_degree"]
    # out-degrees: a2=2, b2=1, c2=1, d2=1 -> [1, 1, 1, 2].
    assert out_deg.min == 1
    assert out_deg.max == 2
    assert out_deg.median == pytest.approx(1.0)
    assert out_deg.p25 == pytest.approx(1.0)
    assert out_deg.p75 == pytest.approx(1.25)

    assert metrics.most_recommended[0]["video_id"] == "a2"
    assert metrics.most_recommended[0]["times_recommended"] == 2
    assert metrics.most_active_sources[0]["video_id"] == "a2"
    assert metrics.most_active_sources[0]["outgoing"] == 2


# ----------------------------------------------------------------------
# Service: temporal
# ----------------------------------------------------------------------
def test_temporal_returns_one_slice_per_requested_run(service) -> None:
    result = service.temporal(["net_r1", "net_r2"])
    assert [slice_model.run_id for slice_model in result.slices] == ["net_r1", "net_r2"]
    assert result.slices[0].node_count == 2
    assert result.slices[0].edge_count == 2
    assert result.slices[1].node_count == 4
    assert result.slices[1].edge_count == 5

    assert len(result.growth) == 1
    growth = result.growth[0]
    assert growth.from_run_id == "net_r1"
    assert growth.to_run_id == "net_r2"
    assert growth.node_growth == 2
    assert growth.edge_growth == 3


def test_temporal_empty_request_returns_empty(service) -> None:
    result = service.temporal([])
    assert result.slices == []
    assert result.growth == []


# ----------------------------------------------------------------------
# Service: edges / export / channels
# ----------------------------------------------------------------------
def test_edges_lists_all_edge_dicts_with_metadata(service) -> None:
    edges = service.edges()
    assert len(edges) == 7
    for edge in edges:
        # Check that all expected attributes exist
        assert hasattr(edge, "source_video_id")
        assert hasattr(edge, "recommended_video_id")
        assert hasattr(edge, "position")
        assert hasattr(edge, "run_id")
        assert hasattr(edge, "title")
        assert hasattr(edge, "channel_id")
        assert hasattr(edge, "thumbnail_url")
        assert hasattr(edge, "views")
        assert hasattr(edge, "likes")
        assert hasattr(edge, "duration")
        assert hasattr(edge, "run_type")
        assert hasattr(edge, "run_name")
        assert hasattr(edge, "source_title")
        assert hasattr(edge, "source_channel_id")
        assert hasattr(edge, "source_channel_name")

        # Check that metadata fields are populated
        assert edge.title is not None
        assert edge.channel_id is not None


def test_edges_run_filter(service) -> None:
    assert len(service.edges(run_id="net_r1")) == 2
    assert len(service.edges(run_id="net_r2")) == 5


def test_edges_channel_filter_matches_source_by_default(service) -> None:
    # Seed video rows so the SOURCE channel resolves (the researcher-facing
    # semantic: show channel X's videos and their 1->N recommendation trees).
    _seed_videos(service._repos)
    assert len(service.edges(channel_id="UC1")) == 2  # a -> b, b -> a
    assert len(service.edges(channel_id="UC2")) == 3  # a2->b2, a2->c2, d2->a2
    assert len(service.edges(channel_id="UC3")) == 2  # b2->c2, c2->a2

    assert len(service.edges(run_id="net_r2", channel_id="UC2")) == 3
    assert len(service.edges(run_id="net_r2", channel_id="UC3")) == 2


def test_edges_channel_filter_target_scope(service) -> None:
    # Legacy/target semantics available via channel_scope="target".
    assert len(service.edges(channel_id="UC1", channel_scope="target")) == 2
    assert len(service.edges(channel_id="UC2", channel_scope="target")) == 3
    assert len(service.edges(channel_id="UC3", channel_scope="target")) == 2


def test_edges_channel_filter_unknown_channel_is_empty(service) -> None:
    _seed_videos(service._repos)
    assert service.edges(channel_id="UC_MISSING") == []


def test_edges_run_and_channel_combined(service) -> None:
    _seed_videos(service._repos)
    assert len(service.edges(run_id="net_r2", channel_id="UC2")) == 3


def test_edges_ordered_by_feed_rank_per_source(service) -> None:
    """Edges are grouped by source and ranked by feed position ascending."""
    edges = service.edges()
    for source in {"a", "b", "a2", "b2", "c2", "d2"}:
        group = [e for e in edges if e.source_video_id == source]
        positions = [e.position for e in group]
        assert positions == sorted(positions), f"{source} not feed-ranked: {positions}"


def test_export_graphml_marker(service) -> None:
    filename, content, media_type = service.export_edges(format="graphml")
    assert filename == "recommendations.graphml"
    assert "<graphml" in content
    assert media_type == "application/xml"


def test_export_edgelist_marker(service) -> None:
    filename, content, _ = service.export_edges(format="edgelist")
    assert filename == "recommendations.edgelist"
    lines = [line for line in content.splitlines() if line]
    assert lines
    for line in lines:
        tokens = line.split()
        # classic edgelist rows are "u v" (optionally followed by data).
        assert len(tokens) >= 2


def test_export_gexf_marker(service) -> None:
    filename, content, media_type = service.export_edges(format="gexf")
    assert filename == "recommendations.gexf"
    assert "<gexf" in content
    assert media_type == "application/gexf+xml"


def test_export_unknown_format_raises_value_error(service) -> None:
    with pytest.raises(ValueError):
        service.export_edges(format="dot")


# ----------------------------------------------------------------------
# Service: export_network (formats, scopes)
# ----------------------------------------------------------------------
def test_export_network_csv_edge_list(service) -> None:
    _seed_videos(service._repos)
    filename, content, media_type = service.export_network(format="csv")
    assert filename == "recommendations.csv"
    assert media_type == "text/csv"
    rows = content.strip().splitlines()
    assert rows[0] == "source,target,weight,relationship_type"
    assert len(rows) == 8  # header + 7 edges
    assert "a,b,1,recommendation" in content
    assert "b,a,1,recommendation" in content


def test_export_network_json_cytoscape_schema(service) -> None:
    _seed_videos(service._repos)
    filename, content, media_type = service.export_network(format="json")
    assert filename == "recommendations.json"
    assert media_type == "application/json"
    payload = json.loads(content)
    assert set(payload) == {"nodes", "links"}
    assert len(payload["nodes"]) == 6
    assert len(payload["links"]) == 7
    node = payload["nodes"][0]["data"]
    assert {"id", "label", "degree", "community_id", "centrality"} <= set(node)
    link = payload["links"][0]["data"]
    assert {"source", "target", "weight", "relationship_type"} <= set(link)
    assert link["relationship_type"] == "recommendation"


def test_export_network_scoped_by_run_ids(service) -> None:
    filename, content, _ = service.export_network(
        format="edgelist", run_ids=["net_r2"]
    )
    assert filename == "recommendations.edgelist"
    lines = [line for line in content.splitlines() if line]
    assert len(lines) == 5
    sources = {line.split()[0] for line in lines}
    assert sources == {"a2", "b2", "c2", "d2"}


def test_export_network_scoped_by_video_ego(service) -> None:
    _, content, _ = service.export_network(
        format="edgelist", video_ids=["a"]
    )
    lines = [line for line in content.splitlines() if line]
    assert len(lines) == 2  # a->b and b->a both touch 'a'


def test_export_network_unknown_format_raises(service) -> None:
    with pytest.raises(ValueError, match="expected one of"):
        service.export_network(format="dot")


def test_export_network_xlsx_workbook(service) -> None:
    """xlsx export returns real workbook bytes with the labeled columns."""
    _seed_videos(service._repos)
    filename, content, media_type = service.export_network(format="xlsx")
    assert filename == "recommendations.xlsx"
    assert media_type.endswith("spreadsheetml.sheet")
    assert isinstance(content, bytes)
    assert content[:2] == b"PK"  # zip magic: a valid OOXML container
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "source_video_id"
    assert len(rows) == 8  # header + 7 edges
    assert any("Title a" in str(cell) for row in rows for cell in row if cell)


def test_export_parity_with_graph_view(service) -> None:
    """Export must contain exactly the nodes/edges the active graph view shows."""
    _seed_videos(service._repos)
    for params in (
        {},
        {"channel_id": "UC1"},
        {"channel_scope": "target"},
        {"connected": "only"},
    ):
        graph = service.graph(**params)
        _, content, _ = service.export_network(format="json", **params)
        payload = json.loads(content)
        graph_node_ids = {n.video_id for n in graph.nodes}
        export_node_ids = {n["data"]["id"] for n in payload["nodes"]}
        assert export_node_ids == graph_node_ids, params
        graph_edges = {(e.source, e.target) for e in graph.edges}
        export_edges = {(l["data"]["source"], l["data"]["target"]) for l in payload["links"]}
        assert export_edges == graph_edges, params


def test_export_channel_projection_parity(service) -> None:
    _seed_videos(service._repos)
    cg = service.channel_graph()
    _, content, _ = service.export_network(format="json", projection="channel")
    payload = json.loads(content)
    assert set(payload) == {"nodes", "links"}
    cg_nodes = {n.channel_id for n in cg.nodes}
    export_nodes = {n["data"]["id"] for n in payload["nodes"]}
    assert export_nodes == cg_nodes
    cg_edges = {(e.source, e.target) for e in cg.edges}
    export_edges = {(l["data"]["source"], l["data"]["target"]) for l in payload["links"]}
    assert export_edges == cg_edges


def test_export_graphml_carries_node_attributes(service) -> None:
    _seed_videos(service._repos)
    filename, content, media_type = service.export_network(format="graphml")
    assert filename == "recommendations.graphml"
    assert "xml" in media_type
    assert "<graphml" in content
    # Spec: node attributes id/label/degree/community_id/centrality present.
    assert "degree" in content
    assert "centrality" in content
    assert "community_id" in content


# ----------------------------------------------------------------------
# Service: merge_networks
# ----------------------------------------------------------------------
def test_merge_identical_scopes_full_overlap(service) -> None:
    scope = NetworkScope(run_id="net_r1")
    result = service.merge_networks(scope, scope)
    assert result.overlap.shared_node_count == 2
    assert result.overlap.shared_edge_count == 2
    assert result.overlap.union_node_count == 2
    assert result.overlap.union_edge_count == 2
    assert result.overlap.nodes_only_in_a == 0
    assert result.overlap.edges_only_in_b == 0
    assert result.overlap.jaccard_node_overlap == 1.0
    assert result.overlap.jaccard_edge_overlap == 1.0
    assert result.merged.node_count == 2
    assert result.merged.edge_count == 2
    assert result.merged.reciprocity == 1.0
    assert result.node_count == 2
    assert result.edge_count == 2


def test_merge_disjoint_run_scopes(service) -> None:
    result = service.merge_networks(
        NetworkScope(run_id="net_r1"), NetworkScope(run_id="net_r2")
    )
    assert result.overlap.shared_node_count == 0
    assert result.overlap.shared_edge_count == 0
    assert result.overlap.union_node_count == 6
    assert result.overlap.union_edge_count == 7
    assert result.overlap.jaccard_node_overlap == 0.0
    assert result.overlap.jaccard_edge_overlap == 0.0


def test_merge_run_against_whole_network_partial_overlap(service) -> None:
    result = service.merge_networks(NetworkScope(run_id="net_r1"), NetworkScope())
    assert result.overlap.scope_a_node_count == 2
    assert result.overlap.shared_node_count == 2
    assert result.overlap.nodes_only_in_a == 0
    assert result.overlap.nodes_only_in_b == 4
    assert result.overlap.union_node_count == 6
    assert result.overlap.union_edge_count == 7
    assert result.overlap.shared_edge_count == 2
    assert result.overlap.edges_only_in_a == 0
    assert result.overlap.edges_only_in_b == 5
    assert result.overlap.jaccard_node_overlap == pytest.approx(2 / 6)
    assert result.overlap.jaccard_edge_overlap == pytest.approx(2 / 7)
    assert result.merged.node_count == 6
    assert result.merged.edge_count == 7
    assert result.edge_count == 7


def test_merge_video_ego_scopes_overlap(service) -> None:
    result = service.merge_networks(
        NetworkScope(video_ids=["a"]), NetworkScope(video_ids=["b"])
    )
    assert result.overlap.scope_a_node_count == 2
    assert result.overlap.scope_b_node_count == 2
    assert result.overlap.shared_node_count == 2
    assert result.overlap.shared_edge_count == 2
    assert result.overlap.union_edge_count == 2
    assert result.overlap.jaccard_node_overlap == 1.0
    assert result.overlap.jaccard_edge_overlap == 1.0


def test_merge_empty_scopes_cover_whole_network(service) -> None:
    result = service.merge_networks(NetworkScope(), NetworkScope())
    assert result.overlap.union_edge_count == 7
    assert result.overlap.jaccard_edge_overlap == 1.0
    assert result.merged.edge_count == 7


def test_merge_labeled_top_degree_nodes(service) -> None:
    _seed_videos(service._repos)
    result = service.merge_networks(NetworkScope(run_id="net_r1"), NetworkScope())
    assert result.merged.top_degree_nodes
    top = result.merged.top_degree_nodes[0]
    assert top.video_id == "a2"
    assert top.title == "Title a2"
    assert top.channel_id == "UC2"
    assert top.total_degree == 4
    by_id = {n.video_id: n for n in result.merged.top_degree_nodes}
    assert by_id["c2"].total_degree == 3
    degrees = [n.total_degree for n in result.merged.top_degree_nodes]
    assert degrees == sorted(degrees, reverse=True)


def test_merge_union_edges_deduplicated(service) -> None:
    """A shared edge is one union edge even when both scopes observe it."""
    result = service.merge_networks(
        NetworkScope(run_id="net_r1"), NetworkScope(run_id="net_r1")
    )
    assert result.edge_count == 2
    pairs = {(e.source, e.target) for e in result.edges}
    assert pairs == {("a", "b"), ("b", "a")}


def test_channel_projection_lists_distinct_channels(service) -> None:
    projection = service.channel_projection()
    assert [f.channel_id for f in projection.channels] == ["UC1", "UC2", "UC3"]
    assert projection.edge_count == 7

    projection = service.channel_projection(run_id="net_r1")
    assert [f.channel_id for f in projection.channels] == ["UC1"]
    assert projection.edge_count == 2


def test_graph_enriched_payload(service) -> None:
    """GET /network/graph returns enriched nodes + edges + facets."""
    _seed_videos(service._repos)
    _seed_runs(service._repos)
    graph = service.graph()
    assert graph.node_count == 6
    assert graph.edge_count == 7

    by_id = {n.video_id: n for n in graph.nodes}
    assert by_id["a"].kind == "both"  # a->b and b->a
    assert by_id["a"].channel_id == "UC1"
    assert by_id["a"].channel_name is not None
    assert by_id["a"].title == "Title a"
    assert by_id["a"].out_degree == 1
    assert by_id["a2"].run_types == ["video"]  # provenance present
    assert by_id["a2"].run_ids == ["net_r2"]
    assert graph.runs  # run facet present
    assert any(f.channel_id == "UC1" for f in graph.channels)
    assert {e.source for e in graph.edges} == {"a", "b", "a2", "b2", "c2", "d2"}


def test_graph_runs_facet_not_clobbered_by_edge_loop(service) -> None:
    """run_id param must survive the edge loop (regression: shadowing bug).

    The old edge loop reused ``run_id`` as a loop variable, so an all-runs
    query (``run_id=None``) was filtered by the last edge's run id, making
    the graph permanently pinned to a single run.
    """
    _seed_videos(service._repos)
    _seed_runs(service._repos)
    all_runs = service.graph()
    assert len(all_runs.runs) == 2
    assert {r["run_id"] for r in all_runs.runs} == {"net_r1", "net_r2"}
    filtered = service.graph(run_id="net_r1")
    assert [r["run_id"] for r in filtered.runs] == ["net_r1"]


def test_graph_channel_filter(service) -> None:
    _seed_videos(service._repos)
    graph = service.graph(channel_id="UC1")
    assert {n.video_id for n in graph.nodes} == {"a", "b"}


def test_graph_falls_back_to_edge_channel_metadata(service) -> None:
    """Unpersisted targets expose provider-observed channel id/name."""
    _seed_videos(service._repos)
    _seed_runs(service._repos)
    repos = service._repos
    repos.recommendations.save_recommendation(
        RecommendationObservation(
            observation_id="r_obs_x",
            collection_run_id="net_r2",
            source_video_id="a",
            recommended_video_id="never_persisted",
            position=0,
            status=RecommendationStatus.OBSERVED,
            channel_id="UC99",
            channel_name="Edge Channel",
            title="Edge Title",
        )
    )
    graph = service.graph()
    by_id = {n.video_id: n for n in graph.nodes}
    node = by_id["never_persisted"]
    assert node.channel_id == "UC99"
    assert node.channel_name == "Edge Channel"
    assert node.title == "Edge Title"
    assert node.kind == "target"

    rows = service.edges()
    row = next(
        r for r in rows if r.recommended_video_id == "never_persisted"
    )
    assert row.channel_id == "UC99"
    assert row.channel_name == "Edge Channel"
    assert row.title == "Edge Title"


# ----------------------------------------------------------------------
# Layer scoping (denormalized layer_index) + channel graph projection
# ----------------------------------------------------------------------
def _stamp_layer(service, run_id: str, layer_index: int) -> None:
    """Re-stamp every edge of a run with ``layer_index``."""
    repos = service._repos
    for edge in repos.recommendations.list_recommendation_edges(run_id=run_id):
        repos.recommendations.save_recommendation(
            edge.model_copy(update={"layer_index": layer_index})
        )


def test_edges_layer_index_filter(service) -> None:
    _stamp_layer(service, "net_r1", 1)
    _stamp_layer(service, "net_r2", 2)
    assert len(service.edges(layer_index=1)) == 2
    assert len(service.edges(layer_index=2)) == 5
    assert all(e.layer_index == 2 for e in service.edges(layer_index=2))


def test_graph_layer_index_filter(service) -> None:
    _stamp_layer(service, "net_r1", 1)
    graph = service.graph(layer_index=1)
    assert graph.edge_count == 2
    assert graph.node_count == 2


def test_graph_isolated_nodes_only(service) -> None:
    """connected=isolated returns only videos with no edge in the slice."""
    _seed_videos(service._repos)
    repos = service._repos
    # A video exists in the corpus but has no recommendation edges.
    repos.videos.upsert_video(
        Video(
            video_id="lonely",
            url="https://www.youtube.com/watch?v=lonely",
            channel_id="UC1",
            title="Lonely",
            first_observed_run_id="net_r1",
        )
    )
    graph = service.graph(connected="isolated")
    assert {n.video_id for n in graph.nodes} == {"lonely"}
    # The connected graph excludes the isolated node.
    full = service.graph()
    assert "lonely" not in {n.video_id for n in full.nodes}


def test_graph_scraped_filter(service) -> None:
    """scraped=scraped / unscraped filter nodes by scrape state."""
    _seed_videos(service._repos)
    repos = service._repos
    repos.videos.mark_recommendations_scraped("a")
    repos.videos.mark_recommendations_scraped("a2")

    scraped_graph = service.graph(scraped="scraped")
    assert {n.video_id for n in scraped_graph.nodes} == {"a", "a2"}
    assert all(n.recommendations_scraped for n in scraped_graph.nodes)

    unscraped_graph = service.graph(scraped="unscraped")
    assert {n.video_id for n in unscraped_graph.nodes} == {"b", "b2", "c2", "d2"}
    assert all(not n.recommendations_scraped for n in unscraped_graph.nodes)


def test_graph_node_carries_recommendations_scraped_flag(service) -> None:
    _seed_videos(service._repos)
    repos = service._repos
    graph = service.graph()
    by_id = {n.video_id: n for n in graph.nodes}
    assert all(by_id[v].recommendations_scraped is False for v in ("a", "b", "a2"))
    repos.videos.mark_recommendations_scraped("a")
    NetworkAnalyticsService.clear_analytics_cache()
    graph = service.graph()
    by_id = {n.video_id: n for n in graph.nodes}
    assert by_id["a"].recommendations_scraped is True
    assert by_id["b"].recommendations_scraped is False


def test_channel_graph_aggregates_weighted_pairs(service) -> None:
    _seed_videos(service._repos)
    projection = service.channel_graph()
    # a->b (UC1->UC1), b->a (UC1->UC1), a2->b2 (UC2->UC3),
    # a2->c2 (UC2->UC3), b2->c2 (UC3->UC3), c2->a2 (UC3->UC2), d2->a2 (UC2->UC2)
    pairs = {
        (e.source, e.target): e.video_edge_count for e in projection.edges
    }
    assert pairs[("UC1", "UC1")] == 2
    assert pairs[("UC2", "UC3")] == 2
    assert pairs[("UC3", "UC3")] == 1
    assert pairs[("UC3", "UC2")] == 1
    assert pairs[("UC2", "UC2")] == 1
    assert projection.node_count == 3
    assert projection.unattributed_edges == 0


def test_channel_graph_layer_scoping(service) -> None:
    _seed_videos(service._repos)
    _stamp_layer(service, "net_r2", 2)
    projection = service.channel_graph(layer_index=2)
    # net_r2 edges only: UC2->UC3 x2, UC3->UC3, UC3->UC2, UC2->UC2 -> 4 pairs.
    assert projection.edge_count == 4
    assert projection.node_count == 2  # UC2 + UC3 (net_r1's UC1 is excluded)
    pairs = {
        (e.source, e.target): e.video_edge_count for e in projection.edges
    }
    assert pairs == {("UC2", "UC3"): 2, ("UC3", "UC3"): 1, ("UC3", "UC2"): 1, ("UC2", "UC2"): 1}


def test_channel_graph_counts_unattributed_edges(service) -> None:
    _seed_videos(service._repos)
    repos = service._repos
    repos.recommendations.save_recommendation(
        RecommendationObservation(
            observation_id="r_obs_nochan",
            collection_run_id="net_r2",
            source_video_id="unknown_source",
            recommended_video_id="unknown_target",
            position=9,
            status=RecommendationStatus.OBSERVED,
            channel_id=None,
        )
    )
    projection = service.channel_graph()
    assert projection.unattributed_edges == 1
    # No synthetic node: dropped edges never create channels.
    assert projection.node_count == 3


def test_graph_empty_network_returns_empty_payload(tmp_path) -> None:
    """No edges at all -> an empty graph (nodes/edges empty, no crash)."""
    repos = build_excel_repositories(
        RepositorySettings(data_dir=str(tmp_path), dataset_name="net_empty")
    )
    empty = NetworkAnalyticsService(repos)
    graph = empty.graph()
    assert graph.node_count == 0
    assert graph.edge_count == 0
    assert graph.nodes == []
    assert graph.edges == []
    assert graph.runs == []
    assert graph.channels == []


def test_graph_self_loop_edge(service) -> None:
    """A video that recommends itself yields one node with in+out degree."""
    repos = service._repos
    repos.recommendations.save_recommendation(
        RecommendationObservation(
            observation_id="r_obs_self",
            collection_run_id="net_r1",
            source_video_id="selfy",
            recommended_video_id="selfy",
            position=0,
            status=RecommendationStatus.OBSERVED,
        )
    )
    repos.videos.upsert_video(
        Video(
            video_id="selfy",
            url="https://www.youtube.com/watch?v=selfy",
            channel_id="UC1",
            title="Self",
            first_observed_run_id="net_r1",
        )
    )
    graph = service.graph()
    by_id = {n.video_id: n for n in graph.nodes}
    assert "selfy" in by_id
    assert by_id["selfy"].in_degree == 1
    assert by_id["selfy"].out_degree == 1
    assert graph.node_count >= 1


def test_graph_parallel_edges_same_pair(service) -> None:
    """Parallel edges between the same pair do not duplicate nodes."""
    repos = service._repos
    for i in range(3):
        repos.recommendations.save_recommendation(
            RecommendationObservation(
                observation_id=f"r_par_{i}",
                collection_run_id="net_r1",
                source_video_id="pa",
                recommended_video_id="pb",
                position=i,
                status=RecommendationStatus.OBSERVED,
            )
        )
    repos.videos.upsert_video(
        Video(
            video_id="pa",
            url="https://www.youtube.com/watch?v=pa",
            channel_id="UC1",
            title="Pa",
            first_observed_run_id="net_r1",
        )
    )
    repos.videos.upsert_video(
        Video(
            video_id="pb",
            url="https://www.youtube.com/watch?v=pb",
            channel_id="UC1",
            title="Pb",
            first_observed_run_id="net_r1",
        )
    )
    graph = service.graph()
    by_id = {n.video_id: n for n in graph.nodes}
    assert by_id["pa"].out_degree == 3
    assert by_id["pb"].in_degree == 3


def test_graph_empty_results_when_filter_matches_nothing(service) -> None:
    """Filters that match no edges/nodes yield an empty payload, not an error."""
    _seed_videos(service._repos)
    repos = service._repos

    empty_run = service.graph(run_id="does-not-exist")
    assert empty_run.node_count == 0
    assert empty_run.edge_count == 0

    empty_layer = service.graph(layer_index=99)
    assert empty_layer.node_count == 0
    assert empty_layer.edge_count == 0

    empty_channel = service.graph(channel_id="UC-NOPE")
    assert empty_channel.node_count == 0
    assert empty_channel.edge_count == 0

    empty_ego = service.graph(video_ids=["zzz-no-such-video"])
    assert empty_ego.node_count == 0

    empty_scraped = service.graph(scraped="scraped")
    assert all(not n.recommendations_scraped for n in empty_scraped.nodes)
    assert empty_scraped.node_count == 0

    empty_isolated = service.graph(connected="isolated")
    assert empty_isolated.node_count == 0


def test_graph_degree_counts_are_consistent(service) -> None:
    """Total out-degree across nodes equals total in-degree across nodes."""
    _seed_videos(service._repos)
    graph = service.graph()
    assert sum(n.out_degree for n in graph.nodes) == sum(n.in_degree for n in graph.nodes)
    assert sum(n.out_degree for n in graph.nodes) == graph.edge_count


# ----------------------------------------------------------------------
# Router endpoints (TestClient)
# ----------------------------------------------------------------------
def test_endpoint_metrics(client) -> None:
    resp = client.get(f"{PREFIX}/network/metrics")
    assert resp.status_code == 200
    body = resp.json()
    assert body["node_count"] == 6
    assert body["edge_count"] == 7
    assert body["is_directed"] is True
    assert 0.0 <= body["density"] <= 1.0


def test_endpoint_graph_advanced_filters(client) -> None:
    """GET /network/graph accepts connected/scraped/layer_index params."""
    resp = client.get(f"{PREFIX}/network/graph")
    assert resp.status_code == 200
    assert resp.json()["node_count"] == 6

    isolated = client.get(
        f"{PREFIX}/network/graph", params={"connected": "isolated"}
    )
    assert isolated.status_code == 200

    bad_connected = client.get(
        f"{PREFIX}/network/graph", params={"connected": "bogus"}
    )
    assert bad_connected.status_code == 400

    bad_scraped = client.get(
        f"{PREFIX}/network/graph", params={"scraped": "bogus"}
    )
    assert bad_scraped.status_code == 400

    layer = client.get(
        f"{PREFIX}/network/graph", params={"layer_index": 0}
    )
    assert layer.status_code == 200


def test_endpoint_network_export(client) -> None:
    resp = client.get(f"{PREFIX}/network/export", params={"format": "graphml"})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/xml")


def test_endpoint_edges_pagination_envelope_with_metadata(client) -> None:
    resp = client.get(f"{PREFIX}/network/edges", params={"page_size": 3})
    assert resp.status_code == 200
    body = resp.json()
    assert set(body) == {"items", "next_cursor", "has_more", "total"}
    assert len(body["items"]) == 3
    assert body["total"] == 7
    assert body["has_more"] is True
    assert body["next_cursor"] is not None
    item = body["items"][0]
    assert set(item) >= {
        "source_video_id",
        "recommended_video_id",
        "position",
        "run_id",
        "title",
        "channel_id",
        "thumbnail_url",
        "views",
        "likes",
        "duration",
        "run_type",
        "run_name",
        "source_title",
        "source_channel_id",
        "source_channel_name",
        "source_thumbnail_url",
    }


def test_endpoint_edges_channel_filter(client) -> None:
    _seed_videos(client.app.state.services["repos"])
    resp = client.get(
        f"{PREFIX}/network/edges",
        params={"channel_id": "UC1", "page_size": 50},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["total"] == 2
    assert {e["source_video_id"] for e in body["items"]} == {"a", "b"}


def test_endpoint_graph(client) -> None:
    _seed_videos(client.app.state.services["repos"])
    _seed_runs(client.app.state.services["repos"])
    resp = client.get(f"{PREFIX}/network/graph")
    assert resp.status_code == 200
    body = resp.json()
    assert body["node_count"] == 6
    assert body["edge_count"] == 7
    assert body["runs"]
    assert any(f["channel_id"] == "UC1" for f in body["channels"])


def test_endpoint_graph_invalid_scope_is_400(client) -> None:
    resp = client.get(f"{PREFIX}/network/graph", params={"channel_scope": "both"})
    assert resp.status_code == 400


def test_endpoint_graph_channel_projection(client) -> None:
    _seed_videos(client.app.state.services["repos"])
    _seed_runs(client.app.state.services["repos"])
    resp = client.get(
        f"{PREFIX}/network/graph", params={"projection": "channel"}
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["projection"] == "channel"
    assert body["node_count"] == 3  # UC1, UC2, UC3
    assert body["edge_count"] == 5
    assert body["runs"]
    assert any(f["channel_id"] == "UC1" for f in body["channels"])


def test_endpoint_graph_invalid_projection_is_400(client) -> None:
    resp = client.get(f"{PREFIX}/network/graph", params={"projection": "text"})
    assert resp.status_code == 400


def test_endpoint_edges_paginates_to_end(client) -> None:
    seen: list[str] = []
    cursor = None
    while True:
        params = {"page_size": 2}
        if cursor:
            params["cursor"] = cursor
        body = client.get(f"{PREFIX}/network/edges", params=params).json()
        seen += [e["source_video_id"] + "->" + e["recommended_video_id"]
                 for e in body["items"]]
        cursor = body["next_cursor"]
        if not body["has_more"]:
            break
    assert len(seen) == 7
    assert len(set(seen)) == 7


def test_endpoint_export_graphml(client) -> None:
    resp = client.get(f"{PREFIX}/network/export", params={"format": "graphml"})
    assert resp.status_code == 200
    assert "<graphml" in resp.text
    assert 'filename="recommendations.graphml"' in resp.headers["content-disposition"]
    assert "xml" in resp.headers["content-type"]


def test_endpoint_export_unknown_format_is_400(client) -> None:
    resp = client.get(f"{PREFIX}/network/export", params={"format": "dot"})
    assert resp.status_code == 400
    assert resp.json()["code"] == "invalid_argument"


def test_endpoint_channels(client) -> None:
    resp = client.get(f"{PREFIX}/network/channels")
    assert resp.status_code == 200
    body = resp.json()
    assert [f["channel_id"] for f in body["channels"]] == ["UC1", "UC2", "UC3"]
    assert body["edge_count"] == 7


# ----------------------------------------------------------------------
# Service: sub-run lineage ("include sub-runs" toggle)
# ----------------------------------------------------------------------
def _seed_run_family(repos) -> None:
    """A parent run with two sub-runs, one of which has a grandchild, plus an
    unrelated run. Each run contributes a single distinct recommendation edge."""
    runs = [
        ("fam_root", None),
        ("fam_child1", "fam_root"),
        ("fam_child2", "fam_root"),
        ("fam_grand", "fam_child1"),
        ("fam_other", None),  # unrelated: must never appear in the family
    ]
    for run_id, parent in runs:
        repos.runs.create_run(
            CollectionRun(
                run_id=run_id,
                run_type=RunType.VIDEO,
                target_url=f"https://www.youtube.com/watch?v={run_id}",
                parent_run_id=parent,
                started_at=utcnow(),
                status="success",
            )
        )
    edges = [
        ("ff_1", "fam_root", "r0", "r1"),
        ("ff_2", "fam_child1", "r1", "r2"),
        ("ff_3", "fam_child2", "r2", "r3"),
        ("ff_4", "fam_grand", "r3", "r4"),
        ("ff_5", "fam_other", "x0", "x1"),
    ]
    for observation_id, run_id, source, target in edges:
        repos.recommendations.save_recommendation(
            RecommendationObservation(
                observation_id=observation_id,
                collection_run_id=run_id,
                source_video_id=source,
                recommended_video_id=target,
                position=0,
                status=RecommendationStatus.OBSERVED,
                channel_id="UCX",
                title=f"T {source}->{target}",
            )
        )


def test_run_family_includes_all_descendants(service) -> None:
    _seed_run_family(service._repos)
    family = service.run_family("fam_root")
    assert set(family) == {"fam_root", "fam_child1", "fam_child2", "fam_grand"}
    assert "fam_other" not in family
    # Cycle safety: re-resolving is idempotent and does not loop forever.
    assert service.run_family("fam_root") == family


def test_graph_with_sub_runs_folds_family(service) -> None:
    _seed_run_family(service._repos)
    family = service.run_family("fam_root")
    g = service.graph(run_ids=family)
    node_ids = {n.video_id for n in g.nodes}
    assert {"r0", "r1", "r2", "r3", "r4"}.issubset(node_ids)
    # Unrelated run is excluded from the family graph.
    assert "x0" not in node_ids and "x1" not in node_ids


def test_graph_single_run_excludes_sub_runs(service) -> None:
    _seed_run_family(service._repos)
    g = service.graph(run_id="fam_root")
    # Without include_sub_runs the parent shows only its own edge.
    assert {n.video_id for n in g.nodes} == {"r0", "r1"}


def test_endpoint_graph_include_sub_runs(family_client) -> None:
    single = family_client.get(
        f"{PREFIX}/network/graph", params={"run_id": "fam_root"}
    ).json()
    assert {n["video_id"] for n in single["nodes"]} == {"r0", "r1"}

    family = family_client.get(
        f"{PREFIX}/network/graph",
        params={"run_id": "fam_root", "include_sub_runs": "true"},
    ).json()
    family_ids = {n["video_id"] for n in family["nodes"]}
    assert {"r0", "r1", "r2", "r3", "r4"}.issubset(family_ids)
    assert "x0" not in family_ids and "x1" not in family_ids


@pytest.fixture
def family_client(tmp_path):
    repo_settings = RepositorySettings(data_dir=str(tmp_path), dataset_name="net")
    repos = build_excel_repositories(repo_settings)
    _seed_recommendations(repos)
    _seed_videos(repos)
    _seed_runs(repos)
    _seed_run_family(repos)
    repos.store.close()
    settings = SocialScienceSettings(
        repository=repo_settings,
        scraper=ScraperSettings(retries=1, retry_backoff=0.0, request_delay_seconds=0),
        collection=CollectionSettings(collect_comments=False),
        api=ApiSettings(prefix=PREFIX),
    )
    app = create_app(settings)
    yield TestClient(app)
