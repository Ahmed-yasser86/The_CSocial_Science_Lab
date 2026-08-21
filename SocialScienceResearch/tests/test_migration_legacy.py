"""Backward-compatible load of legacy workbooks that lack ``observed_at``.

B2 adds observability timestamps to ``transcripts`` and ``recommendations``.
Workbooks written by earlier versions have no such column; opening them must
(a) not crash on the empty-cell datetime conversion and (b) transparently add
the new header so fresh writes remain consistent.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from SocialScienceResearch.config.settings import RepositorySettings
from SocialScienceResearch.persistence.excel_repository import build_excel_repositories

FIXTURES = Path(__file__).parent / "fixtures"
LEGACY = FIXTURES / "legacy_workbook.xlsx"


def test_legacy_workbook_loads_and_columns_are_added(tmp_path) -> None:
    workbook = tmp_path / "dataset.xlsx"
    shutil.copy(LEGACY, workbook)

    rs = RepositorySettings(data_dir=str(tmp_path), dataset_name="dataset")
    repos = build_excel_repositories(rs)

    edges = repos.recommendations.list_recommendation_edges()
    assert len(edges) == 1
    assert edges[0].recommended_video_id == "dst_legacy"
    assert edges[0].observed_at is None  # column added, legacy value empty

    transcripts = repos.transcripts.list_transcripts()
    assert len(transcripts) == 1
    assert transcripts[0].transcript_id == "tx_legacy_1"
    assert transcripts[0].observed_at is None

    # New rows written after the migration carry the header.
    from SocialScienceResearch.domain.models import TranscriptRecord

    repos.transcripts.save_transcript(
        TranscriptRecord(
            transcript_id="tx_new",
            video_id="vid_new",
            collection_run_id="run_new",
            path="transcripts/vid_new.txt",
            lang="en",
            status="available",
        )
    )
    repos.store.close()
    reloaded = build_excel_repositories(rs)
    new_recs = reloaded.transcripts.list_transcripts()
    by_id = {t.transcript_id: t for t in new_recs}
    assert by_id["tx_new"].observed_at is None
    assert "observed_at" in reloaded.store._headers["transcripts"]  # noqa: SLF001
    assert "observed_at" in reloaded.store._headers["recommendations"]  # noqa: SLF001
    reloaded.store.close()


def test_legacy_fixture_missing_observed_at_header(tmp_path) -> None:
    """Guard: the fixture itself must actually be a legacy workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(LEGACY, read_only=True)
    tx = wb["transcripts"]
    assert "observed_at" not in [c.value for c in next(tx.iter_rows())]
    rec = wb["recommendations"]
    assert "observed_at" not in [c.value for c in next(rec.iter_rows())]
    wb.close()