"""One-click Excel export of everything a researcher collected in a Project.

The generic ``POST /export`` endpoint exports a single entity type to a single
sheet. Researchers repeatedly asked for an *easy* way to get **all** the data
they made/collected inside a project as one workbook they can open in Excel for
downstream NetworkX / statistical work. This module builds that multi-sheet
workbook (Videos, Comments, Channels, Recommendations, Runs) from a project's
datasets and samples, gathering provenance (runs) and the recommendation edges
among the project's videos.
"""

from __future__ import annotations

import enum
import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

# Repository wiring is structural; import the concrete class only for typing.
from SocialScienceResearch.persistence.base import Repositories  # noqa: F401


def _write_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    """Append a sheet ``name`` (sheet titles are capped at 31 chars by Excel)."""
    if not rows:
        return
    ws = wb.create_sheet(title=name[:31])
    headers = list(rows[0].keys())
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header)
            if isinstance(value, (dict, list)):
                value = str(value)
            elif isinstance(value, datetime):
                value = value.isoformat()
            elif isinstance(value, date):
                value = value.isoformat()
            elif isinstance(value, enum.Enum):
                value = value.value
            ws.cell(row=row_idx, column=col_idx, value=value)


def _fetch_record(
    repos: Repositories, entity_type: str, member_id: str
) -> dict[str, Any] | None:
    if entity_type == "video":
        rec = repos.videos.get_video(member_id)
    elif entity_type == "comment":
        rec = repos.comments.get_comment(member_id)
    elif entity_type == "channel":
        rec = repos.channels.get_channel(member_id)
    else:
        return None
    return rec.model_dump() if rec is not None else None


def export_project_to_workbook(
    repos: Repositories, project_id: str
) -> tuple[str, bytes]:
    """Return ``(filename, xlsx_bytes)`` with every record a project collected.

    Walks the project's items -> datasets/samples -> members, collecting full
    record dicts per entity type, the recommendation edges among the project's
    videos, and the runs that produced the data (provenance).
    """
    items = repos.project_items.list_items(project_id)

    videos: list[dict[str, Any]] = []
    comments: list[dict[str, Any]] = []
    channels: list[dict[str, Any]] = []
    recommendations: list[dict[str, Any]] = []
    runs_seen: set[str] = set()
    video_ids: set[str] = set()

    def collect(entity_type: str, record: dict[str, Any] | None) -> None:
        if not record:
            return
        if entity_type == "video":
            videos.append(record)
            vid = record.get("video_id")
            if vid:
                video_ids.add(vid)
        elif entity_type == "comment":
            comments.append(record)
        elif entity_type == "channel":
            channels.append(record)
        elif entity_type == "recommendation":
            recommendations.append(record)

    for item in items:
        for dataset_id in item.dataset_ids:
            dataset = repos.datasets.get_dataset(dataset_id)
            if dataset is None:
                continue
            if dataset.created_by_run_id:
                runs_seen.add(dataset.created_by_run_id)
            entity_type = dataset.entity_type
            for member in repos.datasets.list_members(dataset_id):
                collect(entity_type, member)
                if entity_type == "video" and member.get("video_id"):
                    video_ids.add(member["video_id"])

        for sample_id in item.sample_ids:
            sample = repos.samples.get(sample_id)
            if sample is None:
                continue
            entity_type = sample.entity_type
            for member_id in sample.member_ids:
                collect(entity_type, _fetch_record(repos, entity_type, member_id))

    seen_edges: set[tuple[Any, Any, Any]] = set()
    for video_id in video_ids:
        for edge in repos.recommendations.list_recommendation_edges(
            source_video_id=video_id
        ):
            key = (
                edge.source_video_id,
                edge.recommended_video_id,
                edge.collection_run_id,
            )
            if key in seen_edges:
                continue
            seen_edges.add(key)
            if edge.collection_run_id:
                runs_seen.add(edge.collection_run_id)
            recommendations.append(
                {
                    "source_video_id": edge.source_video_id,
                    "recommended_video_id": edge.recommended_video_id,
                    "position": edge.position,
                    "channel_id": edge.channel_id,
                    "title": edge.title,
                    "collection_run_id": edge.collection_run_id,
                }
            )

    run_rows: list[dict[str, Any]] = []
    for run_id in sorted(runs_seen):
        run = repos.runs.get_run(run_id)
        if run is not None:
            run_rows.append(run.model_dump())

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Videos", videos)
    _write_sheet(wb, "Comments", comments)
    _write_sheet(wb, "Channels", channels)
    _write_sheet(wb, "Recommendations", recommendations)
    _write_sheet(wb, "Runs", run_rows)

    if not wb.sheetnames:
        empty = wb.create_sheet("Export")
        empty["A1"] = "No collected data found for this project."

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return f"project_{project_id}_export.xlsx", buffer.getvalue()
