"""Low-level Excel workbook storage.

This module is the *only* place in the application that touches openpyxl. It
provides an append-friendly, header-driven sheet store with:

* stable header rows per sheet (``row 1`` = column names),
* automatic overflow into ``<sheet>__2``, ``<sheet>__3``, ... beyond the
  ``max_rows_per_sheet`` limit (Excel hard limit: 1,048,576 rows/sheet),
* idempotent upserts keyed by a stable column,
* in-memory buffering with periodic auto-flush to disk.

The store is not thread-safe and not designed for concurrent writers (the
Excel backend is intentionally a research-oriented, single-writer
persistence implementation).
"""

from __future__ import annotations

import json
import threading
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from SocialScienceResearch.config.settings import DEFAULT_MAX_ROWS_PER_SHEET
from SocialScienceResearch.utils.logger import get_logger

logger = get_logger(__name__)

MAX_EXCEL_ROWS = 1048576  # hard Excel limit
MAX_CELL_CHARS = 32000  # safety margin below Excel's 32767-char cell limit
_OVERFLOW_SENTINEL = "__CELL_OVERFLOW__"


class WorkbookStore:
    """Header-driven Excel workbook with append, upsert and overflow support.

    Thread-safety: a single re-entrant lock serializes every mutating and
    reading operation so the store is safe under concurrent writers (e.g. the
    multi-worker :class:`JobManager`). Reads take the same lock, so a read
    never observes a half-written cell and can never deadlock against a write
    (``threading.RLock`` re-enters from within a write). This is deliberately
    simple - the workbook is in-memory and research-scale, so serialization
    cost is negligible.
    """

    def __init__(
        self,
        path: str | Path,
        max_rows_per_sheet: int = DEFAULT_MAX_ROWS_PER_SHEET,
        flush_every: int = 1000,
    ) -> None:
        if max_rows_per_sheet > MAX_EXCEL_ROWS - 1:
            raise ValueError(
                f"max_rows_per_sheet must be <= {MAX_EXCEL_ROWS - 1}, got {max_rows_per_sheet}"
            )
        self.path = Path(path)
        self.max_rows_per_sheet = max_rows_per_sheet
        self.flush_every = flush_every

        self._lock = threading.RLock()
        self._headers: dict[str, list[str]] = {}
        self._indexes: dict[tuple[str, str], dict[str, tuple[str, int]]] = {}
        self._since_flush = 0

        # Oversized cell values (exceeding Excel's 32767-char limit) are
        # persisted to per-bucket JSON sidecars keyed by (sheet, header, entity
        # key). Buckets are loaded lazily on demand so boot never pulls the
        # whole sidecar (which can exceed 200MB of raw yt-dlp JSON) into RAM.
        self._overflow_path = self.path.with_name(self.path.stem + ".overflow.json")
        self._overflow: dict[tuple[str, str], dict[str, str]] = {}
        self._load_overflow()

        if self.path.exists():
            self._wb = load_workbook(self.path, read_only=False, data_only=True)
            self._load_headers()
        else:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self._wb = Workbook()
            self._wb.remove(self._wb.active)
        self._close_on_exit = True

    @staticmethod
    def _bucket_filename(sheet: str, header: str) -> str:
        """Deterministic sidecar filename for one (sheet, header) bucket."""
        token = f"{sheet}__{header}".replace("\\", "_").replace("/", "_")
        token = "".join(c if c.isalnum() or c in "-_." else "_" for c in token)
        return f"{token}.overflow.json"

    def _bucket_path(self, sheet: str, header: str) -> Path:
        return self.path.with_name(self._bucket_filename(sheet, header))

    def _load_overflow(self) -> None:
        """Migrate the legacy single-file sidecar into per-bucket files.

        Boot never loads the full sidecar into memory: each (sheet, header)
        bucket lives in its own JSON file and is read on demand. When a legacy
        ``*.overflow.json`` is detected it is split once into buckets; if any
        bucket is unreadable it is skipped (logged) without failing the boot.
        """
        if not self._overflow_path.exists():
            return
        try:
            with open(self._overflow_path, encoding="utf-8") as fh:
                raw = json.load(fh)
        except (ValueError, OSError, MemoryError, RecursionError):
            logger.warning(
                "overflow sidecar %s could not be migrated; oversized cells "
                "will read as empty until the next successful flush",
                self._overflow_path,
            )
            return
        migrated = 0
        for key, bucket in raw.items():
            sheet, sep, header = key.partition("\u0000")
            if not sep or not isinstance(bucket, dict):
                continue
            try:
                with open(self._bucket_path(sheet, header), "w", encoding="utf-8") as fh:
                    json.dump(bucket, fh, ensure_ascii=False)
                migrated += 1
            except OSError as exc:  # noqa: BLE001 - never fail boot on a bucket
                logger.warning(
                    "could not migrate overflow bucket (%s, %s): %s",
                    sheet,
                    header,
                    exc,
                )
        if migrated:
            try:
                self._overflow_path.unlink()
            except OSError:  # pragma: no cover - best-effort cleanup
                pass
        logger.info("migrated %d overflow bucket(s) to per-bucket sidecars", migrated)

    def _bucket(self, sheet: str, header: str) -> dict[str, str]:
        """Return the in-memory bucket for (sheet, header), loading lazily.

        The full overflow is never loaded at boot: a bucket is read from its
        own sidecar file the first time it is touched and cached thereafter.
        """
        key = (sheet, header)
        bucket = self._overflow.get(key)
        if bucket is not None:
            return bucket
        loaded: dict[str, str] = {}
        try:
            with open(self._bucket_path(sheet, header), encoding="utf-8") as fh:
                loaded = json.load(fh)
        except (OSError, ValueError):
            pass
        self._overflow[key] = loaded
        return loaded

    def _value_for_write(
        self, name: str, header: str, key: str, value: Any
    ) -> Any:
        """Route oversized strings to the bucket sidecar; keep cells within Excel limits."""
        if isinstance(value, str) and len(value) > MAX_CELL_CHARS:
            bucket = self._bucket(name, header)
            bucket[str(key)] = value
            return _OVERFLOW_SENTINEL
        return value

    def _value_for_read(
        self, name: str, header: str, key: str, value: Any
    ) -> Any:
        if value == _OVERFLOW_SENTINEL:
            bucket = self._bucket(name, header)
            if key in bucket:
                return bucket[key]
        return value

    # ------------------------------------------------------------------
    # Sheet lifecycle
    # ------------------------------------------------------------------
    def _load_headers(self) -> None:
        for ws in self._wb.worksheets:
            if ws.max_row >= 1:
                headers = [c.value for c in ws[1] if c.value is not None]
                self._headers[ws.title] = headers

    def sheet_names(self) -> list[str]:
        """Return the base (non-overflow) sheet names present in the store."""
        with self._lock:
            names = set(self._headers.keys())
            return sorted(name for name in names if "__" not in name)

    def ensure_sheet(self, name: str, headers: list[str]) -> None:
        """Create the sheet with a header row if it does not exist.

        If the sheet exists, any new headers are appended as additional
        columns (schema evolution without data loss).
        """
        with self._lock:
            if name not in self._wb.sheetnames:
                ws = self._wb.create_sheet(title=name)
                self._headers[name] = []
            ws = self._wb[name]
            existing = self._headers.get(name, [])
            missing = [h for h in headers if h not in existing]
            if not existing:
                for i, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=i, value=header)
                self._headers[name] = list(headers)
            elif missing:
                start = len(existing) + 1
                for j, header in enumerate(missing):
                    ws.cell(row=1, column=start + j, value=header)
                self._headers[name] = existing + missing

    def _data_sheet(self, name: str) -> str:
        """Return the sheet (base or overflow) where the next row is appended."""
        n = 1
        while True:
            candidate = name if n == 1 else f"{name}__{n}"
            if candidate not in self._wb.sheetnames:
                self.ensure_sheet(candidate, self._headers.get(name, []))
            ws = self._wb[candidate]
            if ws.max_row - 1 < self.max_rows_per_sheet:
                return candidate
            n += 1

    # ------------------------------------------------------------------
    # Writes
    # ------------------------------------------------------------------
    def upsert_row(
        self, name: str, key_field: str, headers: list[str], row: dict[str, Any]
    ) -> bool:
        """Insert a row, or update it in place if ``key_field`` already exists.

        Returns ``True`` if the row was newly created, ``False`` if it updated
        an existing row. Idempotent by design: repeated saves do not duplicate.
        """
        with self._lock:
            self.ensure_sheet(name, headers)
            current_headers = self._headers[name]
            key = str(row[key_field])
            index = self._ensure_index(name, key_field)
            if key in index:
                sheet_name, excel_row = index[key]
                ws = self._wb[sheet_name]
                for i, header in enumerate(current_headers, start=1):
                    ws.cell(
                        row=excel_row,
                        column=i,
                        value=self._value_for_write(name, header, key, row.get(header)),
                    )
                self._mark_dirty()
                return False

            sheet_name = self._data_sheet(name)
            ws = self._wb[sheet_name]
            excel_row = ws.max_row + 1
            for i, header in enumerate(current_headers, start=1):
                ws.cell(
                    row=excel_row,
                    column=i,
                    value=self._value_for_write(name, header, key, row.get(header)),
                )
            index[key] = (sheet_name, excel_row)
            self._mark_dirty()
            return True

    def _mark_dirty(self) -> None:
        self._since_flush += 1
        if self._since_flush >= self.flush_every:
            self.save()

    # ------------------------------------------------------------------
    # Reads
    # ------------------------------------------------------------------
    def _ensure_index(
        self, name: str, key_field: str
    ) -> dict[str, tuple[str, int]]:
        index_key = (name, key_field)
        with self._lock:
            if index_key in self._indexes:
                return self._indexes[index_key]
            index: dict[str, tuple[str, int]] = {}
            headers = self._headers.get(name, [])
            key_col = headers.index(key_field) if key_field in headers else None
            n = 1
            while True:
                sheet_name = name if n == 1 else f"{name}__{n}"
                if sheet_name not in self._wb.sheetnames:
                    break
                ws = self._wb[sheet_name]
                for cells in ws.iter_rows(min_row=2):
                    if cells is None:
                        continue
                    if key_col is not None and key_col < len(cells):
                        value = cells[key_col].value
                        if value is not None and str(value) not in index:
                            index[str(value)] = (sheet_name, cells[0].row)
                n += 1
            self._indexes[index_key] = index
            return index

    def read_rows(
        self, name: str, key_field: str | None = None
    ) -> list[dict[str, Any]]:
        """Read all rows of a sheet, concatenating overflow sheets.

        When ``key_field`` is supplied, oversized cells stored in the JSON
        sidecar are resolved back to their full values. Cells for columns the
        row predates (schema evolution) decode as ``None``.
        """
        with self._lock:
            headers = self._headers.get(name, [])
            if not headers:
                return []
            result: list[dict[str, Any]] = []
            n = 1
            while True:
                sheet_name = name if n == 1 else f"{name}__{n}"
                if sheet_name not in self._wb.sheetnames:
                    break
                ws = self._wb[sheet_name]
                for cells in ws.iter_rows(min_row=2, values_only=True):
                    if cells is None:
                        continue
                    key = (
                        str(cells[headers.index(key_field)])
                        if key_field and key_field in headers
                        else None
                    )
                    row = {
                        header: self._value_for_read(
                            name, header, key, cells[i] if i < len(cells) else None
                        )
                        for i, header in enumerate(headers)
                    }
                    if all(v is None or v == "" for v in row.values()):
                        continue
                    result.append(row)
                n += 1
            return result

    def find_row(
        self, name: str, key_field: str, key: str
    ) -> dict[str, Any] | None:
        """Return a single row by its key field, or ``None``."""
        with self._lock:
            location = self._ensure_index(name, key_field).get(str(key))
            if location is None:
                return None
            sheet_name, excel_row = location
            ws = self._wb[sheet_name]
            key = str(key)
            return {
                header: self._value_for_read(
                    name, header, key, ws.cell(row=excel_row, column=i + 1).value
                )
                for i, header in enumerate(self._headers.get(name, []))
            }

    def row_exists(self, name: str, key_field: str, key: str) -> bool:
        return str(key) in self._ensure_index(name, key_field)

    def delete_row(self, name: str, key_field: str, key: str) -> None:
        """Remove a row: blank its cells and drop it from the index.

        Cells are cleared (workbook rows can't be physically removed without
        shifting cells and invalidating every row's position); removing the
        index entry makes the row invisible to ``find_row``/``read_rows``.
        """
        with self._lock:
            index = self._ensure_index(name, key_field)
            location = index.pop(str(key), None)
            if location is None:
                return
            sheet_name, excel_row = location
            ws = self._wb[sheet_name]
            for column in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=column).value = None
            self._mark_dirty()

    # ------------------------------------------------------------------
    # Persistence
    # ------------------------------------------------------------------
    def save(self) -> None:
        """Write the workbook and any oversized-cell sidecars to disk."""
        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self._wb.save(self.path)
            if self._overflow:
                for (sheet, header), bucket in self._overflow.items():
                    try:
                        with open(
                            self._bucket_path(sheet, header), "w", encoding="utf-8"
                        ) as fh:
                            json.dump(bucket, fh, ensure_ascii=False)
                    except OSError as exc:  # noqa: BLE001 - never fail the flush
                        logger.warning(
                            "could not persist overflow bucket (%s, %s): %s",
                            sheet,
                            header,
                            exc,
                        )
            self._since_flush = 0
            logger.debug("saved workbook to %s", self.path)

    def close(self) -> None:
        """Flush any pending data to disk."""
        with self._lock:
            self.save()

    def __enter__(self) -> "WorkbookStore":
        return self

    def __exit__(self, *exc: object) -> None:
        self.close()
